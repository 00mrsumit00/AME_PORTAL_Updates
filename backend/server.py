from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, Query
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.requests import Request
from starlette.responses import FileResponse, StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import io
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import json
import pandas as pd

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import certifi

mongo_url = os.environ.get('MONGO_URL')
db_name = os.environ.get('DB_NAME', 'ame_portal')

# Robust connection for Cloud environments (Render/Atlas)
try:
    masked_url = mongo_url
    if mongo_url and "@" in mongo_url:
        prefix, rest = mongo_url.split("://", 1)
        creds, host = rest.split("@", 1)
        if ":" in creds:
            user, _ = creds.split(":", 1)
            masked_url = f"{prefix}://{user}:****@{host}"
    
    print(f"[DEBUG] Connecting to DB: {db_name}")
    print(f"[DEBUG] Using URI: {masked_url}")
    
    # Sanitize URL: pymongo requires lowercase 'true'/'false' for boolean params
    if mongo_url and "?" in mongo_url:
        base_url, query = mongo_url.split("?", 1)
        params = query.split("&")
        sanitized_params = []
        for p in params:
            if "=" in p:
                k, v = p.split("=", 1)
                if v.lower() in ["true", "false"]:
                    v = v.lower()
                sanitized_params.append(f"{k}={v}")
            else:
                sanitized_params.append(p)
        mongo_url = f"{base_url}?{'&'.join(sanitized_params)}"
    
    client_kwargs = {
        "serverSelectionTimeoutMS": 5000,
        "connectTimeoutMS": 10000,
        "retryWrites": True,
        "tls": True
    }
    
    # Only use certifi if explicitly on a remote host
    if mongo_url and ("localhost" not in mongo_url and "127.0.0.1" not in mongo_url):
        try:
            client_kwargs["tlsCAFile"] = certifi.where()
            client = AsyncIOMotorClient(mongo_url, **client_kwargs)
            # Ping verification removed from global scope to avoid SyntaxError
            db = client[db_name]
            logging.info(f"✅ MongoDB Connected successfully to: {db_name}")
        except Exception as e:
            logging.warning(f"⚠️ Primary connection failed: {e}. Retrying with simple mode...")
            # Fallback: Remove complex SSL params and try again
            client_kwargs.pop("tlsCAFile", None)
            client_kwargs["tlsAllowInvalidCertificates"] = True
            client = AsyncIOMotorClient(mongo_url, **client_kwargs)
            db = client[db_name]
            logging.info(f"✅ MongoDB Connected in fallback mode")
    else:
        client = AsyncIOMotorClient(mongo_url)
        db = client[db_name]
except Exception as e:
    logging.error(f"❌ Critical connection initialization error: {e}")
    client = AsyncIOMotorClient(mongo_url)
    db = client[db_name]

JWT_SECRET = os.environ.get('JWT_SECRET', 'fallback-secret')
JWT_ALGORITHM = "HS256"
JWT_EXPIRY_HOURS = 24 * 365  # Keep user logged in for 1 year until they manually log out

UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

DOCUMENT_TYPES = {
    "MEDICAL": ["NEET Scorecard", "Allotment Letter", "10th Marksheet", "12th Marksheet", "Aadhaar Card", "Photograph", "Category Certificate"],
    "ENGINEERING": ["CET Scorecard", "Allotment Letter", "10th Marksheet", "12th Marksheet", "Aadhaar Card", "Photograph", "Category Certificate"],
    "DSE": ["Diploma Marksheets", "Allotment Letter", "10th Marksheet", "Aadhaar Card", "Photograph", "Category Certificate"]
}

MAHARASHTRA_DISTRICTS = [
    "Mumbai", "Pune", "Nagpur", "Thane", "Nashik", "Chhatrapati Sambhajinagar", "Solapur",
    "Kolhapur", "Sangli", "Satara", "Ratnagiri", "Sindhudurg", "Raigad", "Palghar",
    "Ahmednagar", "Jalgaon", "Dhule", "Nandurbar", "Beed", "Latur", "Dharashiv",
    "Nanded", "Parbhani", "Hingoli", "Washim", "Akola", "Amravati", "Yavatmal",
    "Buldhana", "Wardha", "Chandrapur", "Gadchiroli", "Gondia", "Bhandara", "Jalna"
]

CATEGORIES = ["GEN", "OBC", "SC", "ST", "EWS", "NT-B", "NT-C", "NT-D", "SEBC", "SBC", "VJ-DT(A)"]

SPECIAL_RESERVATIONS = ["DEF-1", "DEF-2", "DEF-3", "MKB", "Hilly Area", "PWD", "Orphan", "None"]

MEDICAL_COUNSELLING_TYPES = {
    "State Quota": ["MBBS", "BDS", "BAMS", "BUMS", "BHMS", "BPTH", "B.O.Th", "BASLP", "BPO", "B.Sc Nursing"],
    "AIIMS": [],
    "AFMC": [],
    "JIPMER": [],
    "AIQ 15% GOV": ["MBBS", "BDS", "BAMS", "BHMS"],
    "Deemed": ["MBBS", "BDS", "BAMS"],
    "Other State": ["MBBS", "BAMS"]
}

# ---- CUTOFF DATA ----
CUTOFF_DATA_FILE = ROOT_DIR / "data" / "NEET.xlsx"
cached_cutoff_sheets = {}
cutoff_metadata = {"sheets": [], "filters": {}}

# ---- COLLEGE DATA (legacy NEET.xlsx) ----
colleges_df = None

# ---- COLLEGE DETAILS (new process-based file) ----
COLLEGE_DETAILS_FILE = ROOT_DIR / "data" / "COLLEGE DETAILS.xlsx"
college_process_sheets = {}  # { sheet_name: DataFrame }

ENGG_COLLEGE_DETAILS_FILE = ROOT_DIR / "data" / "ENGG College Details.xlsx"
engg_college_process_sheets = {}  # { sheet_name: DataFrame }

def load_cutoff_data():
    global cached_cutoff_sheets, cutoff_metadata
    if not CUTOFF_DATA_FILE.exists():
        logger.warning(f"Cutoff data file not found at {CUTOFF_DATA_FILE}")
        return
    
    try:
        logger.info(f"Loading cutoff data from {CUTOFF_DATA_FILE}...")
        xl = pd.ExcelFile(CUTOFF_DATA_FILE)
        cutoff_metadata["sheets"] = xl.sheet_names
        for sheet_name in xl.sheet_names:
            # header=None to preserve original rows for list-of-lists conversion
            df = xl.parse(sheet_name, header=None)
            cached_cutoff_sheets[sheet_name] = df
            
            # Identify State column for filters (search Row 0)
            row0 = df.iloc[0].astype(str).str.upper().tolist()
            states = []
            if 'STATE' in row0:
                idx = row0.index('STATE')
                # Use dropna() and unique() to get clean list of states from index 2 onwards
                states = sorted(df.iloc[2:, idx].dropna().unique().astype(str).tolist())
            
            cutoff_metadata["filters"][sheet_name] = {
                "states": states
            }
        logger.info(f"Cutoff data loaded successfully ({len(xl.sheet_names)} sheets).")
    except Exception as e:
        logger.error(f"Failed to load cutoff data: {e}")

def load_colleges_data():
    global colleges_df
    if not CUTOFF_DATA_FILE.exists():
        logger.warning("NEET.xlsx not found — cannot load college data.")
        return
    try:
        xl = pd.ExcelFile(CUTOFF_DATA_FILE)
        if "COLLEGE DETAILS" not in xl.sheet_names:
            logger.warning("Sheet 'COLLEGE DETAILS' not found in NEET.xlsx")
            return
        raw_df = xl.parse("COLLEGE DETAILS", header=None)
        ncols = raw_df.shape[1]
        data_df = raw_df.iloc[2:].copy()

        # Map columns based on count (handle 7, 8, 9, 10 col variants)
        if ncols >= 10:
            data_df = data_df.iloc[:, :10]
            data_df.columns = ["code", "course_type", "name", "status",
                               "establishment", "fees", "intake_capacity",
                               "hostel", "website", "contact_no"]
        elif ncols >= 8:
            data_df = data_df.iloc[:, :8]
            data_df.columns = ["code", "course_type", "name", "status",
                               "establishment", "fees", "intake_capacity", "hostel"]
            data_df["website"] = ""
            data_df["contact_no"] = ""
        else:
            data_df = data_df.iloc[:, :7]
            data_df.columns = ["code", "name", "status",
                               "establishment", "fees", "intake_capacity", "hostel"]
            data_df["course_type"] = "MBBS"
            data_df["website"] = ""
            data_df["contact_no"] = ""

        data_df = data_df.fillna("")

        # Normalize status values so duplicates collapse into one
        STATUS_MAP = {
            "government": "Govt.",
            "govt": "Govt.",
            "private": "Private",
            "government aided": "Govt. Aided",
            "govt aided": "Govt. Aided",
            "govt. aided": "Govt. Aided",
            "semi-govt.": "Semi-Govt.",
            "semi govt": "Semi-Govt.",
            "municipal": "Municipal",
            "deemed": "Deemed",
        }
        def normalize_status(s):
            s = str(s).strip()
            return STATUS_MAP.get(s.lower(), s)
        data_df["status"] = data_df["status"].apply(normalize_status)

        def safe_int(x):
            try:
                return str(int(float(x))) if x != "" else ""
            except Exception:
                return str(x)
        data_df["code"] = data_df["code"].apply(safe_int)
        data_df["establishment"] = data_df["establishment"].apply(safe_int)
        data_df["intake_capacity"] = data_df["intake_capacity"].apply(safe_int)
        data_df = data_df.reset_index(drop=True)
        data_df = data_df[data_df["name"].astype(str).str.strip() != ""]
        colleges_df = data_df
        print(f"College Details loaded: {len(colleges_df)} colleges")
        logger.info(f"College Details loaded: {len(colleges_df)} colleges")
    except Exception as e:
        logger.error(f"Failed to load college data: {e}")

def load_college_details():
    """Load COLLEGE DETAILS.xlsx — multi-sheet process-based file."""
    global college_process_sheets
    if not COLLEGE_DETAILS_FILE.exists():
        logger.warning(f"COLLEGE DETAILS.xlsx not found at {COLLEGE_DETAILS_FILE}")
        return
    try:
        STATUS_MAP = {
            "government": "Govt.", "govt": "Govt.", "private": "Private",
            "government aided": "Govt. Aided", "govt aided": "Govt. Aided",
            "govt. aided": "Govt. Aided", "semi-govt.": "Semi-Govt.",
            "semi govt": "Semi-Govt.", "municipal": "Municipal", "deemed": "Deemed",
        }
        def norm_status(s):
            s = str(s).strip()
            return STATUS_MAP.get(s.lower(), s)
        def safe_int(x):
            try: return str(int(float(x))) if str(x).strip() not in ('', 'nan') else ''
            except: return str(x)
        def safe_str(x):
            s = str(x).strip()
            return '' if s in ('nan', 'None', 'NaN') else s

        xl = pd.ExcelFile(COLLEGE_DETAILS_FILE)
        result = {}
        for sheet in xl.sheet_names:
            raw = xl.parse(sheet, header=None)
            if len(raw) < 2: continue
            
            # Identify columns from header row (row 0)
            headers = [str(h).strip() for h in raw.iloc[0]]
            df = raw.iloc[1:].copy()
            
            # Map based on headers
            col_map = {}
            for i, h in enumerate(headers):
                h_upper = str(h).upper()
                if 'CODE' in h_upper: col_map['code'] = i
                elif 'SR NO' in h_upper or 'SR. NO' in h_upper and 'code' not in col_map: col_map['code'] = i
                elif 'QUOTA' in h_upper: col_map['quota'] = i
                elif 'COURSE' in h_upper: col_map['course_type'] = i
                elif 'COLLEGE' in h_upper and 'NAME' not in h_upper and 'WEBSITE' not in h_upper and 'CONTACT' not in h_upper and 'COLLEGES IN MAHARASHTRA' == h_upper: col_map['name'] = i
                elif 'NAME' in h_upper or 'COLLEGES' in h_upper or 'COLLEGE' in h_upper and 'WEBSITE' not in h_upper and 'CONTACT' not in h_upper: col_map['name'] = i
                elif 'STATUS' in h_upper: col_map['status'] = i
                elif 'STATE' in h_upper: col_map['state'] = i
                elif 'ESTABLISHMENT' in h_upper: col_map['establishment'] = i
                elif 'FEE' in h_upper: col_map['fees'] = i
                elif 'INTAKE' in h_upper: col_map['intake_capacity'] = i
                elif 'HOSTEL' in h_upper: col_map['hostel'] = i
                elif 'WEBSITE' in h_upper: col_map['website'] = i
                elif 'CONTACT' in h_upper: col_map['contact_no'] = i

            # Ensure all keys exist in df
            final_df = pd.DataFrame()
            for key in ['code', 'quota', 'course_type', 'name', 'status', 'state', 
                        'establishment', 'fees', 'intake_capacity', 'hostel', 'website', 'contact_no']:
                if key in col_map:
                    final_df[key] = df.iloc[:, col_map[key]]
                else:
                    final_df[key] = ''

            df = final_df.fillna('')
            df['status'] = df['status'].apply(norm_status)
            # Default status for AIIMS if empty
            if sheet == 'AIIMS':
                df.loc[df['status'] == '', 'status'] = 'Govt.'
            
            df['establishment'] = df['establishment'].apply(safe_int)
            df['intake_capacity'] = df['intake_capacity'].apply(safe_int)
            df['fees'] = df['fees'].apply(safe_str)
            df['website'] = df['website'].apply(safe_str)
            df['contact_no'] = df['contact_no'].apply(safe_str)
            
            def safe_state(x):
                val = safe_str(x)
                return val if val else 'All India'
            df['state'] = df['state'].apply(safe_state)
            
            df = df[df['name'].astype(str).str.strip() != '']
            df = df.reset_index(drop=True)
            result[sheet] = df
            print(f"Process '{sheet}' loaded: {len(df)} colleges")
            logger.info(f"Process '{sheet}' loaded: {len(df)} colleges")
        college_process_sheets = result
    except Exception as e:
        logger.error(f"Failed to load COLLEGE DETAILS.xlsx: {e}")

def load_engg_college_details():
    """Load ENGG College Details.xlsx — handles column name variations per sheet."""
    global engg_college_process_sheets
    if not ENGG_COLLEGE_DETAILS_FILE.exists():
        logger.warning(f"ENGG College Details.xlsx not found at {ENGG_COLLEGE_DETAILS_FILE}")
        return
    try:
        def safe_int(x):
            try: return str(int(float(x))) if str(x).strip() not in ('', 'nan') else ''
            except: return str(x)
        def safe_str(x):
            s = str(x).strip()
            return '' if s in ('nan', 'None', 'NaN') else s
        def get_col(df, *candidates):
            """Return first matching column from df, or empty Series."""
            for c in candidates:
                if c in df.columns:
                    return df[c]
            return pd.Series([''] * len(df), index=df.index)

        xl = pd.ExcelFile(ENGG_COLLEGE_DETAILS_FILE)
        result = {}
        for sheet in xl.sheet_names:
            raw = xl.parse(sheet, header=None)
            if len(raw) < 2: continue

            headers = [str(h).strip().upper() for h in raw.iloc[0]]
            df = raw.iloc[1:].copy()
            df.columns = headers

            final_df = pd.DataFrame()
            # code — IIT/IIIT/NIT/GFTIS use "INST CODE"; MAHARASHTRA uses "CODE"
            final_df['code'] = get_col(df, 'INST CODE', 'CODE')
            final_df['quota'] = get_col(df, 'ADMISSION QUOTA')
            # name — each sheet uses a different column heading for the college name
            final_df['name'] = get_col(df,
                'COLLEGES IN MAHARASHTRA',           # IIT, MAHARASHTRA, IIIT sheets
                'NATIONAL INSTITUTE OF TECHNOLOGY (NIT)',  # NIT sheet
                'GOVERNMENT FUNDED TECHNICAL INSTITUTE',   # GFTIS sheet
                'COLLEGE NAME', 'NAME'               # generic fallback
            )
            # establishment — MAHARASHTRA has typo "ESTABLISMENT IN"
            final_df['establishment'] = get_col(df, 'ESTABLISHMENT IN', 'ESTABLISMENT IN', 'ESTABLISHMENT')
            final_df['fees'] = get_col(df, 'FEES')
            final_df['intake_capacity'] = get_col(df, 'INTAKE CAPACITY', 'INTAKE')
            # hostel — consistent typo "HOSTEL AVAILABELITY" across all sheets
            final_df['hostel'] = get_col(df, 'HOSTEL AVAILABELITY', 'HOSTEL AVAILABILITY', 'HOSTEL')
            final_df['website'] = get_col(df, 'COLLEGE WEBSITE', 'WEBSITE')
            # contact — MAHARASHTRA uses "CONTACT NUMBER"; others use "COLLEGE CONTACT NUMBER"
            final_df['contact_no'] = get_col(df, 'COLLEGE CONTACT NUMBER', 'CONTACT NUMBER', 'CONTACT NO')
            # status — only MAHARASHTRA sheet has this column
            final_df['status'] = get_col(df, 'STATUS')

            final_df = final_df.fillna('')
            final_df['establishment'] = final_df['establishment'].apply(safe_int)
            final_df['intake_capacity'] = final_df['intake_capacity'].apply(safe_int)
            final_df['fees'] = final_df['fees'].apply(safe_str)
            final_df['website'] = final_df['website'].apply(safe_str)
            final_df['contact_no'] = final_df['contact_no'].apply(safe_str)

            final_df = final_df[final_df['name'].astype(str).str.strip() != '']
            final_df = final_df.reset_index(drop=True)
            result[sheet] = final_df
            print(f"Engg Process '{sheet}' loaded: {len(final_df)} colleges")
            logger.info(f"Engg Process '{sheet}' loaded: {len(final_df)} colleges")
        engg_college_process_sheets = result
    except Exception as e:
        logger.error(f"Failed to load ENGG College Details.xlsx: {e}")

OTHER_STATE_MBBS_STATES = ["Telangana", "Andhra Pradesh", "Karnataka", "Uttar Pradesh"]
OTHER_STATE_BAMS_STATES = ["Karnataka", "Madhya Pradesh"]

app = FastAPI()

# CORS is configured at the end of the file

api_router = APIRouter() # Prefix moved to include_router for reliability

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

load_cutoff_data()
load_colleges_data()
load_college_details()
load_engg_college_details()

@app.get("/health")
async def health_check():
    return {"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat()}

@api_router.get("/health")
async def api_health_check():
    db_status = "unknown"
    try:
        # Check DB connection
        await client.admin.command('ping')
        db_status = "connected"
    except Exception as e:
        db_status = f"error: {str(e)}"
    
    return {
        "status": "ok",
        "api_prefix": "/api",
        "database": db_status,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

# ==================== CUTOFF ROUTES ====================

@api_router.get("/cutoff/metadata")
async def get_cutoff_metadata():
    return cutoff_metadata

@api_router.post("/cutoff/data")
async def get_cutoff_data(req: Dict[str, Any]):
    sheet_name = req.get("sheetName")
    category = req.get("category", "").upper()
    state = req.get("state", "").upper()
    
    logger.info(f"📊 Request for cutoff data: sheet={sheet_name}, category={category}, state={state}")
    
    if not sheet_name or sheet_name not in cached_cutoff_sheets:
        raise HTTPException(status_code=400, detail="Invalid sheet name")
        
    df = cached_cutoff_sheets[sheet_name]
    row0 = df.iloc[0].astype(str).str.upper().tolist()
    
    # Base masks
    mask = pd.Series([True] * len(df))
    
    # Filter by Category if present
    if 'CATEGORY' in row0:
        cat_idx = row0.index('CATEGORY')
        if category and category != 'ALL CATEGORY':
            mask &= (df.iloc[:, cat_idx].astype(str).str.upper() == category)
            
    # Filter by State if present
    if 'STATE' in row0:
        state_idx = row0.index('STATE')
        if state and state != 'ALL STATE':
            mask &= (df.iloc[:, state_idx].astype(str).str.upper() == state)
            
    # Keep header rows (0 and 1)
    header_indices = [0, 1]
    
    # Filtered data (skipping headers for filtering)
    data_indices = df.index[mask].tolist()
    # Remove 0 and 1 if they are in data_indices to avoid duplicates
    data_indices = [i for i in data_indices if i > 1]
    
    # Combined result (headers + top 100 results)
    final_indices = header_indices + data_indices[:100]
    result_df = df.loc[final_indices]
    
    # Convert to list and sanitize for JSON (handling NaN/Inf)
    raw_list = result_df.values.tolist()
    import math
    def sanitize(val):
        if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
            return None
        return val
    
    result_list = [[sanitize(cell) for cell in row] for row in raw_list]
    
    return {"data": result_list}

# ==================== COLLEGE ROUTES ====================

@api_router.get("/colleges/metadata")
async def get_colleges_metadata():
    if colleges_df is None:
        return {"statuses": [], "hostel_options": [], "course_types": []}
    statuses = sorted([s for s in colleges_df["status"].unique().tolist() if s != ""])
    hostel_options = sorted([h for h in colleges_df["hostel"].unique().tolist() if h != ""])
    course_types = sorted([c for c in colleges_df["course_type"].unique().tolist() if c != ""])
    return {"statuses": statuses, "hostel_options": hostel_options, "course_types": course_types}

@api_router.get("/colleges/search")
async def search_colleges(
    search: str = Query(""),
    status: str = Query(""),        # comma-separated: "Govt.,Govt. Aided"
    hostel: str = Query(""),
    course_type: str = Query("")
):
    if colleges_df is None:
        return []
    import math
    mask = pd.Series([True] * len(colleges_df))
    if search.strip():
        mask &= colleges_df["name"].str.contains(search.strip(), case=False, na=False)
    if status.strip() and status.strip() != "ALL":
        status_list = [s.strip() for s in status.split(",") if s.strip()]
        if status_list:
            mask &= colleges_df["status"].isin(status_list)
    if hostel.strip() and hostel.strip() != "ALL":
        mask &= (colleges_df["hostel"] == hostel.strip())
    if course_type.strip() and course_type.strip() != "ALL":
        mask &= (colleges_df["course_type"].str.upper() == course_type.strip().upper())
    result = colleges_df[mask]
    records = result.to_dict(orient="records")
    def clean(v):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return v
    return [{k: clean(v) for k, v in row.items()} for row in records]

# ==================== COLLEGE PROCESS ROUTES ====================

@api_router.get("/colleges/processes")
async def get_processes():
    """Return the list of available process/quota sheets."""
    return {"processes": list(college_process_sheets.keys())}

@api_router.get("/colleges/process-metadata")
async def get_process_metadata(process: str = Query("State Quota")):
    """Return dynamic filter options for the given process."""
    if process not in college_process_sheets:
        return {"statuses": [], "course_types": [], "hostel_options": [], "states": []}
    df = college_process_sheets[process]
    statuses = sorted([s for s in df['status'].unique() if s])
    course_types = sorted([c for c in df['course_type'].unique() if c])
    hostel_opts = sorted([h for h in df['hostel'].unique() if h])
    # Extract state
    states = []
    if process in ['AIIMS', 'DEEMED']:
        states = sorted([s for s in df['state'].unique() if s])
    
    return {
        "statuses": statuses,
        "course_types": course_types,
        "hostel_options": hostel_opts,
        "states": states,
    }

@api_router.get("/colleges/process-search")
async def process_search_colleges(
    process: str = Query("State Quota"),
    search: str = Query(""),
    course_type: str = Query(""),
    status: str = Query(""),        # comma-separated for multi-select
    hostel: str = Query(""),
    state: str = Query(""),         # AIIMS state filter
    fees_min: int = Query(0),
    fees_max: int = Query(0),
):
    import math
    if process not in college_process_sheets:
        return []
    df = college_process_sheets[process]
    mask = pd.Series([True] * len(df))
    if search.strip():
        mask &= df['name'].str.contains(search.strip(), case=False, na=False)
    if course_type.strip() and course_type.strip() != 'ALL':
        mask &= df['course_type'].str.upper() == course_type.strip().upper()
    if status.strip() and status.strip() != 'ALL':
        status_list = [s.strip() for s in status.split(',') if s.strip()]
        if status_list:
            mask &= df['status'].isin(status_list)
    if hostel.strip() and hostel.strip() != 'ALL':
        mask &= df['hostel'] == hostel.strip()
    if state.strip() and state.strip().upper() not in ['ALL', 'ALL STATES']:
        mask &= df['state'].str.lower() == state.strip().lower()
    if (fees_min > 0 or fees_max > 0) and process != 'AIIMS':
        def parse_fees(v):
            try: return int(float(str(v).replace(',','')))
            except: return 0
        fees_nums = df['fees'].apply(parse_fees)
        if fees_min > 0: mask &= fees_nums >= fees_min
        if fees_max > 0: mask &= fees_nums <= fees_max
    result = df[mask]
    records = result.to_dict(orient='records')
    def clean(v):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)): return None
        return v
    return [{k: clean(v) for k, v in row.items()} for row in records]

@api_router.get("/engg-colleges/processes")
async def get_engg_processes():
    return {"processes": list(engg_college_process_sheets.keys())}

@api_router.get("/engg-colleges/process-metadata")
async def get_engg_process_metadata(process: str = Query("MAHARASHTRA")):
    if process not in engg_college_process_sheets:
        return {"statuses": [], "course_types": [], "hostel_options": [], "states": []}
    df = engg_college_process_sheets[process]
    
    statuses = sorted([str(s) for s in df['status'].unique() if str(s).strip() != '']) if 'status' in df.columns else []
    course_types = sorted([str(c) for c in df['course_type'].unique() if str(c).strip() != '']) if 'course_type' in df.columns else []
    hostel_opts = sorted([str(s) for s in df['hostel'].unique() if str(s).strip() != '']) if 'hostel' in df.columns else []
    states = sorted([str(s) for s in df['state'].unique() if str(s).strip() != '']) if 'state' in df.columns else []
    
    return {
        "statuses": statuses,
        "course_types": course_types,
        "hostel_options": hostel_opts,
        "states": states,
    }

@api_router.get("/engg-colleges/process-search")
async def process_search_engg_colleges(
    process: str = Query("MAHARASHTRA"),
    search: str = Query(""),
    course_type: str = Query(""),
    status: str = Query(""),        # comma-separated for multi-select
    hostel: str = Query(""),
    state: str = Query(""),
    fees_min: int = Query(0),
    fees_max: int = Query(0),
):
    import math
    if process not in engg_college_process_sheets:
        return []
    df = engg_college_process_sheets[process].copy().reset_index(drop=True)
    mask = pd.Series([True] * len(df), index=df.index)
    
    if search.strip():
        mask &= df['name'].str.contains(search.strip(), case=False, na=False, regex=False)
        
    if course_type.strip() and course_type.strip() != 'ALL' and 'course_type' in df.columns:
        mask &= df['course_type'].str.upper() == course_type.strip().upper()
        
    if status.strip() and 'status' in df.columns:
        status_list = [s.strip().upper() for s in status.split(',') if s.strip()]
        if status_list:
            mask &= df['status'].astype(str).str.upper().str.strip().isin(status_list)
            
    if hostel.strip() and hostel.strip() != 'ALL' and 'hostel' in df.columns:
        mask &= df['hostel'].str.upper() == hostel.strip().upper()
        
    if state.strip() and state.strip() != 'ALL' and 'state' in df.columns:
        mask &= df['state'].str.upper() == state.strip().upper()
    
    if fees_min > 0 or fees_max > 0:
        def parse_fees(v):
            v_str = str(v).lower().replace(',','').replace('rs.','').strip()
            if not v_str: return 0
            multiplier = 1
            if 'lakh' in v_str or 'lac' in v_str:
                multiplier = 100000
                v_str = v_str.replace('lakh','').replace('lac','').strip()
            try: return int(float(v_str) * multiplier)
            except: return 0
            
        fees_nums = df['fees'].apply(parse_fees)
        if fees_min > 0: mask &= fees_nums >= fees_min
        if fees_max > 0: mask &= fees_nums <= fees_max
        
    result = df[mask]
    records = result.to_dict(orient='records')
    def clean(v):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)): return None
        return v
    return [{k: clean(v) for k, v in row.items()} for row in records]

# ==================== MODELS ====================

class PublicUserCreate(BaseModel):
    full_name: str
    phone: str
    password: str

class PublicUserLogin(BaseModel):
    phone: str
    password: str

class PublicProfileUpdate(BaseModel):
    full_name: Optional[str] = None
    neet_score: Optional[int] = None
    category: Optional[str] = None
    gender: Optional[str] = None
    district: Optional[str] = None
    special_reservation: Optional[str] = None

class LoginRequest(BaseModel):
    email: str
    password: str

class BranchCreate(BaseModel):
    name: str
    location: str
    contact_number: str
    head_name: str
    head_email: str
    head_password: str

class BranchUpdate(BaseModel):
    name: Optional[str] = None
    location: Optional[str] = None
    contact_number: Optional[str] = None
    is_active: Optional[bool] = None

import csv
import io

class StudentRegistration(BaseModel):
    full_name: str
    phone: str
    parent_phone: str = ""
    gender: str = ""
    dob: str = ""
    district: str = ""
    category: str = ""
    caste: str = ""
    special_reservation: str = "None"
    fresher_repeater: str = "Fresher"
    counselling_type: str
    medical_selections: List[Dict[str, Any]] = []
    selection_display: str = ""
    engineering_type: str = ""
    branch_priority: str = ""
    registered_by: str = ""
    registered_by_id: str = ""
    payment_amount: float = 0
    pending_amount: float = 0
    concession_amount: float = 0
    payment_mode: str = "CASH"
    utr_number: str = ""
    payment_note: str = ""

class BulkImportSummary(BaseModel):
    success_count: int
    failed_count: int
    errors: List[str] = []

class StudentAdditionalUpdate(BaseModel):
    mother_name: str = ""
    email: str = ""
    religion: str = ""
    family_income: str = ""
    aadhaar_number: str = ""
    address: str = ""
    taluka: str = ""
    address_district: str = ""
    pin_code: str = ""
    edu_10_year: str = ""
    edu_10_place: str = ""
    edu_10_school_type: str = ""
    edu_10_board: str = ""
    edu_10_state: str = ""
    edu_10_district: str = ""
    edu_10_roll: str = ""
    edu_10_total: str = ""
    edu_10_obtained: str = ""
    edu_10_percentage: str = ""
    edu_10_school: str = ""
    edu_10_school_address: str = ""
    edu_10_pin: str = ""
    edu_11_year: str = ""
    edu_11_place: str = ""
    edu_11_college_type: str = ""
    edu_11_board: str = ""
    edu_11_state: str = ""
    edu_11_district: str = ""
    edu_11_roll: str = ""
    edu_11_total: str = ""
    edu_11_obtained: str = ""
    edu_11_percentage: str = ""
    edu_11_college: str = ""
    edu_11_college_address: str = ""
    edu_11_pin: str = ""
    edu_12_year: str = ""
    edu_12_place: str = ""
    edu_12_college_type: str = ""
    edu_12_board: str = ""
    edu_12_state: str = ""
    edu_12_district: str = ""
    edu_12_roll: str = ""
    edu_12_total: str = ""
    edu_12_obtained: str = ""
    edu_12_percentage: str = ""
    edu_12_college: str = ""
    edu_12_college_address: str = ""
    edu_12_pin: str = ""
    father_occupation: str = ""
    father_qualification: str = ""
    mother_occupation: str = ""
    mother_qualification: str = ""

class StaffCreate(BaseModel):
    name: str
    email: str
    password: str

class ChatMessageCreate(BaseModel):
    content: str = ""
    attachment_url: str = ""

class VerificationDecision(BaseModel):
    field_name: str
    decision: str
    comment: str = ""

class VerificationSubmit(BaseModel):
    decisions: List[VerificationDecision]

class RegistrationUpdate(BaseModel):
    full_name: Optional[str] = None
    phone: Optional[str] = None
    parent_phone: Optional[str] = None
    gender: Optional[str] = None
    dob: Optional[str] = None
    district: Optional[str] = None
    category: Optional[str] = None
    caste: Optional[str] = None
    special_reservation: Optional[str] = None
    fresher_repeater: Optional[str] = None
    counselling_type: Optional[str] = None
    medical_selections: Optional[List[Dict[str, Any]]] = None
    selection_display: Optional[str] = None
    engineering_type: Optional[str] = None
    branch_priority: Optional[str] = None
    registered_by: Optional[str] = None
    registered_by_id: Optional[str] = None
    payment_amount: Optional[float] = None
    pending_amount: Optional[float] = None
    concession_amount: Optional[float] = None
    payment_mode: Optional[str] = None
    utr_number: Optional[str] = None
    payment_note: Optional[str] = None
    updated_by_id: Optional[str] = None
    edit_reason: Optional[str] = ""

# ==================== AUTH UTILS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, role: str, branch_id: str = None) -> str:
    payload = {
        "user_id": user_id,
        "role": role,
        "branch_id": branch_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRY_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = auth_header.split(" ")[1]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user or not user.get("is_active", True):
            raise HTTPException(status_code=401, detail="User not found or inactive")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_public_user(request: Request) -> dict:
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = auth_header.split(" ")[1]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("role") != "PUBLIC":
            raise HTTPException(status_code=403, detail="Not a public user")
        user = await db.public_users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def require_role(*roles):
    async def dependency(request: Request):
        user = await get_current_user(request)
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return dependency

async def generate_reg_no(counselling_type: str, branch_id: str) -> str:
    prefix = {"MEDICAL": "MED", "ENGINEERING": "ENGG", "DSE": "DSE"}.get(counselling_type, "STU")
    counter_name = f"reg_{prefix.lower()}_{branch_id}"
    counter = await db.counters.find_one_and_update(
        {"name": counter_name},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    return f"{prefix} {counter['seq']:04d}"

async def log_audit(user_id: str, branch_id: str, action: str, target_id: str, description: str):
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "branch_id": branch_id or "",
        "action": action,
        "target_id": target_id,
        "description": description,
        "created_at": datetime.now(timezone.utc).isoformat()
    })

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/login")
async def login(req: LoginRequest):
    user = await db.users.find_one({"email": req.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="Account deactivated")
    if not verify_password(req.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_token(user["id"], user["role"], user.get("branch_id"))
    return {
        "token": token,
        "user": {
            "id": user["id"], "email": user["email"], "name": user["name"],
            "role": user["role"], "branch_id": user.get("branch_id")
        }
    }

@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    return {
        "id": user["id"], "email": user["email"], "name": user["name"],
        "role": user["role"], "branch_id": user.get("branch_id")
    }

# ==================== PUBLIC USER ROUTES ====================

@api_router.post("/public/auth/register")
async def public_register(data: PublicUserCreate):
    existing = await db.public_users.find_one({"phone": data.phone})
    if existing:
        raise HTTPException(status_code=400, detail="Phone number already registered")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "full_name": data.full_name,
        "phone": data.phone,
        "password": hash_password(data.password),
        "role": "PUBLIC",
        "neet_score": None,
        "category": None,
        "gender": None,
        "district": None,
        "special_reservation": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.public_users.insert_one(user_doc)
    
    token = create_token(user_id, "PUBLIC")
    return {
        "token": token,
        "user": {
            "id": user_id, "full_name": data.full_name, "phone": data.phone,
            "role": "PUBLIC", "neet_score": None, "category": None, 
            "gender": None, "district": None, "special_reservation": None
        }
    }

@api_router.post("/public/auth/login")
async def public_login(data: PublicUserLogin):
    user = await db.public_users.find_one({"phone": data.phone}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["id"], "PUBLIC")
    return {
        "token": token,
        "user": {
            "id": user["id"], "full_name": user.get("full_name"), "phone": user["phone"],
            "role": "PUBLIC", "neet_score": user.get("neet_score"), 
            "category": user.get("category"), "gender": user.get("gender"),
            "district": user.get("district"), "special_reservation": user.get("special_reservation")
        }
    }

@api_router.get("/public/profile")
async def get_public_profile(user: dict = Depends(get_public_user)):
    return user

@api_router.put("/public/profile")
async def update_public_profile(data: PublicProfileUpdate, user: dict = Depends(get_public_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        return user
    await db.public_users.update_one({"id": user["id"]}, {"$set": update_data})
    updated_user = await db.public_users.find_one({"id": user["id"]}, {"_id": 0})
    return updated_user


# ==================== PUBLIC ADMIN ROUTES ====================

# --- Pydantic Models for Public Admin ---
class PublicAdminLogin(BaseModel):
    email: str
    password: str

class PublicSubAdminCreate(BaseModel):
    name: str
    email: str
    password: str

class PublicAdminPasswordChange(BaseModel):
    user_id: str
    new_password: str

class PublicAdminUserToggle(BaseModel):
    user_id: str
    is_active: bool

class PublicBroadcastMessage(BaseModel):
    message: str
    title: Optional[str] = "Announcement"

class PublicChatMessage(BaseModel):
    student_id: str
    message: str

class PortalTickerUpdate(BaseModel):
    text: str
    link: Optional[str] = ""

class PublicUpdateCreate(BaseModel):
    type: str # 'OFFICIAL_UPDATE' or 'UPCOMING_EVENT'
    title: str
    description: Optional[str] = ""
    link: str
    date: str
    is_active: Optional[bool] = True

# --- Auth helper for public admin ---
async def get_public_admin(request: Request) -> dict:
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = auth_header.split(" ")[1]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("role") not in ("PUBLIC_ADMIN", "PUBLIC_SUBADMIN"):
            raise HTTPException(status_code=403, detail="Not a public admin")
        admin = await db.public_admins.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not admin or not admin.get("is_active", True):
            raise HTTPException(status_code=401, detail="Admin not found or inactive")
        return admin
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_public_master_admin(request: Request) -> dict:
    admin = await get_public_admin(request)
    if admin.get("role") != "PUBLIC_ADMIN":
        raise HTTPException(status_code=403, detail="Master admin access required")
    return admin

async def log_public_audit(admin_id: str, admin_name: str, action: str, target_id: str, description: str):
    await db.public_audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "admin_id": admin_id,
        "admin_name": admin_name,
        "action": action,
        "target_id": target_id,
        "description": description,
        "created_at": datetime.now(timezone.utc).isoformat()
    })

# --- Public Admin Auth ---
@api_router.post("/public-admin/auth/login")
async def public_admin_login(data: PublicAdminLogin):
    admin = await db.public_admins.find_one({"email": data.email}, {"_id": 0})
    if not admin:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if not admin.get("is_active", True):
        raise HTTPException(status_code=401, detail="Account deactivated")
    if not verify_password(data.password, admin["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    token = create_token(admin["id"], admin["role"])
    return {
        "token": token,
        "user": {
            "id": admin["id"], "name": admin["name"], "email": admin["email"],
            "role": admin["role"]
        }
    }

@api_router.get("/public-admin/auth/me")
async def public_admin_me(admin: dict = Depends(get_public_admin)):
    return {
        "id": admin["id"], "name": admin["name"], "email": admin["email"],
        "role": admin["role"]
    }

# --- Public Admin Dashboard ---
@api_router.get("/public-admin/dashboard")
async def public_admin_dashboard(admin: dict = Depends(get_public_admin)):
    total_users = await db.public_users.count_documents({})
    active_users = await db.public_users.count_documents({"is_active": {"$ne": False}})
    disabled_users = await db.public_users.count_documents({"is_active": False})

    # Gender distribution
    gender_pipeline = [
        {"$match": {"gender": {"$ne": None}}},
        {"$group": {"_id": "$gender", "count": {"$sum": 1}}}
    ]
    gender_data = {}
    async for doc in db.public_users.aggregate(gender_pipeline):
        gender_data[doc["_id"]] = doc["count"]

    # Category distribution
    category_pipeline = [
        {"$match": {"category": {"$ne": None}}},
        {"$group": {"_id": "$category", "count": {"$sum": 1}}}
    ]
    category_data = {}
    async for doc in db.public_users.aggregate(category_pipeline):
        category_data[doc["_id"]] = doc["count"]

    # District distribution
    district_pipeline = [
        {"$match": {"district": {"$ne": None}}},
        {"$group": {"_id": "$district", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 15}
    ]
    district_data = {}
    async for doc in db.public_users.aggregate(district_pipeline):
        district_data[doc["_id"]] = doc["count"]

    # NEET Score distribution (buckets)
    score_pipeline = [
        {"$match": {"neet_score": {"$ne": None, "$gt": 0}}},
        {"$bucket": {
            "groupBy": "$neet_score",
            "boundaries": [0, 200, 300, 400, 500, 600, 700, 721],
            "default": "Other",
            "output": {"count": {"$sum": 1}}
        }}
    ]
    score_data = []
    try:
        async for doc in db.public_users.aggregate(score_pipeline):
            label = f"{doc['_id']}" if doc['_id'] != "Other" else "Other"
            score_data.append({"range": label, "count": doc["count"]})
    except Exception:
        score_data = []

    # Sub-admin count
    sub_admin_count = await db.public_admins.count_documents({"role": "PUBLIC_SUBADMIN"})

    # Recent registrations (last 7 days)
    seven_days_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    recent_count = await db.public_users.count_documents({"created_at": {"$gte": seven_days_ago}})

    return {
        "total_users": total_users,
        "active_users": active_users,
        "disabled_users": disabled_users,
        "recent_registrations": recent_count,
        "sub_admin_count": sub_admin_count,
        "gender_distribution": gender_data,
        "category_distribution": category_data,
        "district_distribution": district_data,
        "score_distribution": score_data
    }

# --- User Management ---
@api_router.get("/public-admin/users")
async def public_admin_list_users(
    search: str = Query(""),
    page: int = Query(1),
    limit: int = Query(50),
    admin: dict = Depends(get_public_admin)
):
    query = {}
    if search:
        query = {"$or": [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}}
        ]}

    total = await db.public_users.count_documents(query)
    skip = (page - 1) * limit
    users = []
    cursor = db.public_users.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit)
    async for user in cursor:
        user.pop("password", None)
        users.append(user)

    return {"users": users, "total": total, "page": page, "limit": limit}

@api_router.put("/public-admin/users/password")
async def public_admin_change_password(data: PublicAdminPasswordChange, admin: dict = Depends(get_public_admin)):
    user = await db.public_users.find_one({"id": data.user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    hashed = hash_password(data.new_password)
    await db.public_users.update_one({"id": data.user_id}, {"$set": {"password": hashed}})
    await log_public_audit(admin["id"], admin["name"], "CHANGE_PASSWORD", data.user_id, f"Changed password for {user.get('full_name', 'Unknown')}")
    return {"message": "Password changed successfully"}

@api_router.put("/public-admin/users/toggle")
async def public_admin_toggle_user(data: PublicAdminUserToggle, admin: dict = Depends(get_public_admin)):
    user = await db.public_users.find_one({"id": data.user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    await db.public_users.update_one({"id": data.user_id}, {"$set": {"is_active": data.is_active}})
    status = "enabled" if data.is_active else "disabled"
    await log_public_audit(admin["id"], admin["name"], "TOGGLE_USER", data.user_id, f"{status.capitalize()} user {user.get('full_name', 'Unknown')}")
    return {"message": f"User {status} successfully"}

@api_router.delete("/public-admin/users/{user_id}")
async def public_admin_delete_user(user_id: str, admin: dict = Depends(get_public_admin)):
    user = await db.public_users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    await db.public_users.delete_one({"id": user_id})
    await log_public_audit(admin["id"], admin["name"], "DELETE_USER", user_id, f"Deleted user {user.get('full_name', 'Unknown')}")
    return {"message": "User deleted successfully"}

# --- Sub-Admin Management (Master Admin Only) ---
@api_router.get("/public-admin/sub-admins")
async def list_sub_admins(admin: dict = Depends(require_public_master_admin)):
    sub_admins = []
    cursor = db.public_admins.find({"role": "PUBLIC_SUBADMIN"}, {"_id": 0})
    async for sa in cursor:
        sa.pop("password", None)
        sub_admins.append(sa)
    return sub_admins

@api_router.post("/public-admin/sub-admins")
async def create_sub_admin(data: PublicSubAdminCreate, admin: dict = Depends(require_public_master_admin)):
    existing_email = await db.public_admins.find_one({"email": data.email})
    if existing_email:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    existing_phone = await db.public_admins.find_one({"phone": data.phone})
    if existing_phone:
        raise HTTPException(status_code=400, detail="Phone number already registered")

    sa_id = str(uuid.uuid4())
    sa_doc = {
        "id": sa_id,
        "name": data.name,
        "email": data.email,
        "phone": data.phone,
        "password": hash_password(data.password),
        "role": "PUBLIC_SUBADMIN",
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.public_admins.insert_one(sa_doc)
    await log_public_audit(admin["id"], admin["name"], "CREATE_SUBADMIN", sa_id, f"Created sub-admin {data.name} ({data.email})")
    return {"id": sa_id, "name": data.name, "email": data.email, "phone": data.phone, "role": "PUBLIC_SUBADMIN", "is_active": True}

@api_router.delete("/public-admin/sub-admins/{sa_id}")
async def delete_sub_admin(sa_id: str, admin: dict = Depends(require_public_master_admin)):
    sa = await db.public_admins.find_one({"id": sa_id, "role": "PUBLIC_SUBADMIN"})
    if not sa:
        raise HTTPException(status_code=404, detail="Sub-admin not found")
    await db.public_admins.delete_one({"id": sa_id})
    await log_public_audit(admin["id"], admin["name"], "DELETE_SUBADMIN", sa_id, f"Deleted sub-admin {sa.get('name', 'Unknown')}")
    return {"message": "Sub-admin deleted"}

@api_router.put("/public-admin/sub-admins/{sa_id}/toggle")
async def toggle_sub_admin(sa_id: str, admin: dict = Depends(require_public_master_admin)):
    sa = await db.public_admins.find_one({"id": sa_id, "role": "PUBLIC_SUBADMIN"})
    if not sa:
        raise HTTPException(status_code=404, detail="Sub-admin not found")
    new_status = not sa.get("is_active", True)
    await db.public_admins.update_one({"id": sa_id}, {"$set": {"is_active": new_status}})
    await log_public_audit(admin["id"], admin["name"], "TOGGLE_SUBADMIN", sa_id, f"{'Enabled' if new_status else 'Disabled'} sub-admin {sa.get('name', 'Unknown')}")
    return {"message": f"Sub-admin {'enabled' if new_status else 'disabled'}"}

# --- Chat System ---
@api_router.get("/public-admin/chats")
async def admin_get_chat_list(admin: dict = Depends(get_public_admin)):
    """Get list of all students with their latest chat message"""
    pipeline = [
        {"$sort": {"created_at": -1}},
        {"$group": {
            "_id": "$student_id",
            "last_message": {"$first": "$message"},
            "last_sender": {"$first": "$sender_role"},
            "last_time": {"$first": "$created_at"},
            "unread_count": {"$sum": {"$cond": [
                {"$and": [{"$eq": ["$sender_role", "STUDENT"]}, {"$eq": ["$read", False]}]},
                1, 0
            ]}}
        }},
        {"$sort": {"last_time": -1}}
    ]
    chat_list = []
    async for doc in db.public_chats.aggregate(pipeline):
        student = await db.public_users.find_one({"id": doc["_id"]}, {"_id": 0, "full_name": 1, "phone": 1})
        if student:
            chat_list.append({
                "student_id": doc["_id"],
                "student_name": student.get("full_name", "Unknown"),
                "student_phone": student.get("phone", ""),
                "last_message": doc["last_message"],
                "last_sender": doc["last_sender"],
                "last_time": doc["last_time"],
                "unread_count": doc["unread_count"]
            })
    return chat_list

@api_router.get("/public-admin/chats/{student_id}")
async def admin_get_chat_messages(student_id: str, admin: dict = Depends(get_public_admin)):
    """Get all messages for a specific student"""
    messages = []
    cursor = db.public_chats.find({"student_id": student_id}, {"_id": 0}).sort("created_at", 1)
    async for msg in cursor:
        messages.append(msg)
    # Mark all student messages as read
    await db.public_chats.update_many(
        {"student_id": student_id, "sender_role": "STUDENT", "read": False},
        {"$set": {"read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    return messages

@api_router.post("/public-admin/chats/send")
async def admin_send_message(data: PublicChatMessage, admin: dict = Depends(get_public_admin)):
    msg_doc = {
        "id": str(uuid.uuid4()),
        "student_id": data.student_id,
        "sender_id": admin["id"],
        "sender_name": admin["name"],
        "sender_role": "ADMIN",
        "message": data.message,
        "read": False,
        "delivered": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.public_chats.insert_one(msg_doc)
    return {"message": "Sent", "id": msg_doc["id"]}

# Student-side chat endpoints
@api_router.get("/public/chats")
async def student_get_chats(user: dict = Depends(get_public_user)):
    messages = []
    cursor = db.public_chats.find({"student_id": user["id"]}, {"_id": 0}).sort("created_at", 1)
    async for msg in cursor:
        messages.append(msg)
    # Mark admin messages as read
    await db.public_chats.update_many(
        {"student_id": user["id"], "sender_role": "ADMIN", "read": False},
        {"$set": {"read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    return messages

@api_router.post("/public/chats/send")
async def student_send_message(data: dict, user: dict = Depends(get_public_user)):
    msg_doc = {
        "id": str(uuid.uuid4()),
        "student_id": user["id"],
        "sender_id": user["id"],
        "sender_name": user.get("full_name", "Student"),
        "sender_role": "STUDENT",
        "message": data.get("message", ""),
        "read": False,
        "delivered": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.public_chats.insert_one(msg_doc)
    return {"message": "Sent", "id": msg_doc["id"]}

@api_router.get("/public/chats/unread-count")
async def student_unread_count(user: dict = Depends(get_public_user)):
    count = await db.public_chats.count_documents({
        "student_id": user["id"], "sender_role": "ADMIN", "read": False
    })
    return {"unread": count}

# --- Portal Configuration (Ticker) ---
@api_router.get("/public/ticker")
async def get_portal_ticker():
    config = await db.portal_config.find_one({"type": "ticker"}, {"_id": 0})
    if not config:
        return {"text": "", "link": ""}
    return config

@api_router.post("/public-admin/ticker")
async def update_portal_ticker(data: PortalTickerUpdate, admin: dict = Depends(get_public_admin)):
    await db.portal_config.update_one(
        {"type": "ticker"},
        {"$set": {"text": data.text, "link": data.link, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    await log_public_audit(admin["id"], admin["name"], "UPDATE_TICKER", "TICKER", f"Updated ticker text: {data.text[:30]}...")
    return {"message": "Ticker updated"}

# --- Portal Updates (PDFs & Events) ---
@api_router.get("/public/portal-updates")
async def student_get_portal_updates():
    updates = []
    cursor = db.public_updates.find({"is_active": True}, {"_id": 0}).sort("created_at", -1)
    async for u in cursor:
        updates.append(u)
    return updates

@api_router.get("/public-admin/portal-updates")
async def admin_list_portal_updates(admin: dict = Depends(get_public_admin)):
    updates = []
    cursor = db.public_updates.find({}, {"_id": 0}).sort("created_at", -1)
    async for u in cursor:
        updates.append(u)
    return updates

@api_router.post("/public-admin/portal-updates")
async def admin_create_portal_update(data: PublicUpdateCreate, admin: dict = Depends(get_public_admin)):
    try:
        u_id = str(uuid.uuid4())
        doc = {
            "id": u_id,
            "type": data.type,
            "title": data.title,
            "description": data.description,
            "link": data.link,
            "date": data.date,
            "is_active": data.is_active,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": admin.get("id", "unknown")
        }
        await db.public_updates.insert_one(doc)
        try:
            await log_public_audit(admin.get("id", "unknown"), admin.get("name", "Admin"), "CREATE_UPDATE", u_id, f"Created {data.type}: {data.title}")
        except Exception as audit_err:
            logging.error(f"Audit log failed: {audit_err}")
            
        return {"message": "Update published", "id": u_id}
    except Exception as e:
        logging.error(f"Error creating portal update: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@api_router.put("/public-admin/portal-updates/{u_id}")
async def admin_update_portal_update(u_id: str, data: PublicUpdateCreate, admin: dict = Depends(get_public_admin)):
    try:
        res = await db.public_updates.update_one(
            {"id": u_id},
            {"$set": {
                "type": data.type,
                "title": data.title,
                "description": data.description,
                "link": data.link,
                "date": data.date,
                "is_active": data.is_active,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        if res.matched_count == 0:
            raise HTTPException(status_code=404, detail="Update not found")
            
        try:
            await log_public_audit(admin.get("id", "unknown"), admin.get("name", "Admin"), "EDIT_UPDATE", u_id, f"Updated {data.type}: {data.title}")
        except Exception as audit_err:
            logging.error(f"Audit log failed: {audit_err}")

        return {"message": "Update updated"}
    except Exception as e:
        logging.error(f"Error updating portal update: {e}")
        raise HTTPException(status_code=500, detail=f"Update error: {str(e)}")

@api_router.delete("/public-admin/portal-updates/{u_id}")
async def admin_delete_portal_update(u_id: str, admin: dict = Depends(get_public_admin)):
    res = await db.public_updates.delete_one({"id": u_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Update not found")
    await log_public_audit(admin["id"], admin["name"], "DELETE_UPDATE", u_id, "Deleted portal update")
    return {"message": "Update deleted"}

@api_router.get("/public/updates")
async def student_get_updates(user: dict = Depends(get_public_user)):
    """Get global announcements for students"""
    broadcasts = []
    cursor = db.public_broadcasts.find({}, {"_id": 0}).sort("created_at", -1).limit(20)
    async for b in cursor:
        broadcasts.append(b)
    return broadcasts

# --- Broadcast ---
@api_router.post("/public-admin/broadcast")
async def admin_broadcast(data: PublicBroadcastMessage, admin: dict = Depends(get_public_admin)):
    """Send a broadcast message to ALL students"""
    all_users = []
    cursor = db.public_users.find({"is_active": {"$ne": False}}, {"id": 1, "_id": 0})
    async for u in cursor:
        all_users.append(u["id"])

    # Insert broadcast into each student's chat
    broadcast_docs = []
    now = datetime.now(timezone.utc).isoformat()
    for uid in all_users:
        broadcast_docs.append({
            "id": str(uuid.uuid4()),
            "student_id": uid,
            "sender_id": admin["id"],
            "sender_name": admin["name"],
            "sender_role": "ADMIN",
            "message": data.message,
            "is_broadcast": True,
            "broadcast_title": data.title,
            "read": False,
            "delivered": True,
            "created_at": now
        })

    if broadcast_docs:
        await db.public_chats.insert_many(broadcast_docs)

    # Also store in broadcast history
    await db.public_broadcasts.insert_one({
        "id": str(uuid.uuid4()),
        "admin_id": admin["id"],
        "admin_name": admin["name"],
        "title": data.title,
        "message": data.message,
        "recipients": len(all_users),
        "created_at": now
    })

    await log_public_audit(admin["id"], admin["name"], "BROADCAST", "", f"Sent broadcast to {len(all_users)} students: {data.title}")
    return {"message": f"Broadcast sent to {len(all_users)} students"}

@api_router.get("/public-admin/broadcasts")
async def admin_list_broadcasts(admin: dict = Depends(get_public_admin)):
    broadcasts = []
    cursor = db.public_broadcasts.find({}, {"_id": 0}).sort("created_at", -1).limit(50)
    async for b in cursor:
        broadcasts.append(b)
    return broadcasts

# --- Audit Logs ---
@api_router.get("/public-admin/audit-logs")
async def admin_get_audit_logs(
    page: int = Query(1),
    limit: int = Query(50),
    admin: dict = Depends(get_public_admin)
):
    total = await db.public_audit_logs.count_documents({})
    skip = (page - 1) * limit
    logs = []
    cursor = db.public_audit_logs.find({}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit)
    async for log in cursor:
        logs.append(log)
    return {"logs": logs, "total": total}

# --- Export Users to Excel ---
@api_router.get("/public-admin/export-users")
async def export_users_excel(admin: dict = Depends(get_public_admin)):
    users = []
    cursor = db.public_users.find({}, {"_id": 0, "password": 0})
    async for user in cursor:
        users.append(user)

    if not users:
        raise HTTPException(status_code=404, detail="No users to export")

    df = pd.DataFrame(users)
    # Reorder columns
    cols = ["full_name", "phone", "gender", "category", "district", "neet_score", "special_reservation", "created_at"]
    existing_cols = [c for c in cols if c in df.columns]
    df = df[existing_cols + [c for c in df.columns if c not in existing_cols and c != "id"]]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Public Users")
    output.seek(0)

    await log_public_audit(admin["id"], admin["name"], "EXPORT_DATA", "", f"Exported {len(users)} users to Excel")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=public_users_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"}
    )



@api_router.get("/admin/dashboard")
async def admin_dashboard(
    date_from: str = Query(""), date_to: str = Query(""),
    branch_id: str = Query(""), counselling_type: str = Query(""),
    user: dict = Depends(require_role("ADMIN"))
):
    query = {}
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    if branch_id:
        query["branch_id"] = branch_id
    if counselling_type:
        query["counselling_type"] = counselling_type
    active_branches = await db.branches.count_documents({"is_active": True})
    total_students = await db.student_profiles.count_documents(query)
    verified = await db.student_profiles.count_documents({**query, "verification_status": "VERIFIED"})
    # Revenue with filters
    pay_match = {}
    if query:
        student_ids = [s["id"] async for s in db.student_profiles.find(query, {"id": 1})]
        if student_ids:
            pay_match = {"student_id": {"$in": student_ids}}
        else:
            pay_match = {"student_id": "__none__"}
    rev_result = await db.payments.aggregate([{"$match": pay_match}, {"$group": {"_id": None, "total": {"$sum": "$amount"}}}]).to_list(1)
    revenue = rev_result[0]["total"] if rev_result else 0
    
    # Total Expense across all branches
    exp_match = {"is_deleted": False}
    if date_from:
        exp_match["date"] = {"$gte": date_from[:10]}
    if date_to:
        exp_match.setdefault("date", {})["$lte"] = date_to[:10]
    if branch_id:
        exp_match["branch_id"] = branch_id

    exp_result = await db.branch_expenses.aggregate([{"$match": exp_match}, {"$group": {"_id": None, "total": {"$sum": "$amount"}}}]).to_list(1)
    total_expense = exp_result[0]["total"] if exp_result else 0

    return {
        "active_branches": active_branches, "total_students": total_students,
        "verified_enrollments": verified, "gross_revenue": revenue,
        "total_expense": total_expense
    }

@api_router.get("/admin/branches")
async def list_branches(user: dict = Depends(require_role("ADMIN"))):
    branches = await db.branches.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for b in branches:
        b["student_count"] = await db.student_profiles.count_documents({"branch_id": b["id"]})
        b["staff_count"] = await db.users.count_documents({"branch_id": b["id"], "role": {"$in": ["BRANCH_HEAD", "STAFF"]}})
    return branches

@api_router.post("/admin/branches")
async def create_branch(data: BranchCreate, user: dict = Depends(require_role("ADMIN"))):
    existing = await db.users.find_one({"email": data.head_email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already in use")
    branch_id = str(uuid.uuid4())
    branch = {
        "id": branch_id, "name": data.name, "location": data.location,
        "contact_number": data.contact_number, "is_active": True,
        "head_name": data.head_name, "head_email": data.head_email,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.branches.insert_one(branch)
    head_user = {
        "id": str(uuid.uuid4()), "email": data.head_email,
        "password": hash_password(data.head_password), "name": data.head_name,
        "role": "BRANCH_HEAD", "branch_id": branch_id, "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(head_user)
    await log_audit(user["id"], "", "CREATE_BRANCH", branch_id, f"Created branch: {data.name}")
    branch.pop("_id", None)
    return branch

@api_router.put("/admin/branches/{branch_id}")
async def update_branch(branch_id: str, data: BranchUpdate, user: dict = Depends(require_role("ADMIN"))):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    result = await db.branches.update_one({"id": branch_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Branch not found")
    if "is_active" in update_data:
        await db.users.update_many({"branch_id": branch_id}, {"$set": {"is_active": update_data["is_active"]}})
    await log_audit(user["id"], "", "UPDATE_BRANCH", branch_id, f"Updated branch: {branch_id}")
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    return branch

@api_router.delete("/admin/branches/{branch_id}")
async def delete_branch(branch_id: str, user: dict = Depends(require_role("ADMIN"))):
    branch = await db.branches.find_one({"id": branch_id})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    student_ids = [s["id"] async for s in db.student_profiles.find({"branch_id": branch_id}, {"id": 1})]
    if student_ids:
        await db.documents.delete_many({"student_id": {"$in": student_ids}})
        await db.payments.delete_many({"student_id": {"$in": student_ids}})
        await db.verification_logs.delete_many({"student_id": {"$in": student_ids}})
    await db.student_profiles.delete_many({"branch_id": branch_id})
    await db.users.delete_many({"branch_id": branch_id})
    await db.audit_logs.delete_many({"branch_id": branch_id})
    await db.branches.delete_one({"id": branch_id})
    await log_audit(user["id"], "", "DELETE_BRANCH", branch_id, f"Deleted branch: {branch.get('name', '')}")
    return {"message": "Branch deleted"}

@api_router.get("/admin/analytics")
async def admin_analytics(
    date_from: str = Query(""), date_to: str = Query(""),
    branch_id: str = Query(""), counselling_type: str = Query(""),
    user: dict = Depends(require_role("ADMIN"))
):
    match = {}
    if date_from:
        match["created_at"] = {"$gte": date_from}
    if date_to:
        match.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    if branch_id:
        match["branch_id"] = branch_id
    if counselling_type:
        match["counselling_type"] = counselling_type

    base = [{"$match": match}] if match else []

    type_dist = await db.student_profiles.aggregate(base + [{"$group": {"_id": "$counselling_type", "count": {"$sum": 1}}}]).to_list(10)
    status_dist = await db.student_profiles.aggregate(base + [{"$group": {"_id": "$verification_status", "count": {"$sum": 1}}}]).to_list(10)
    cat_dist = await db.student_profiles.aggregate(base + [{"$group": {"_id": "$category", "count": {"$sum": 1}}}]).to_list(20)
    location_dist = await db.student_profiles.aggregate(base + [{"$group": {"_id": "$district", "count": {"$sum": 1}}}, {"$sort": {"count": -1}}, {"$limit": 20}]).to_list(20)

    # Score distribution (12th percentage ranges)
    score_pipeline = base + [
        {"$match": {"education_12th.percentage": {"$exists": True, "$ne": ""}}},
        {"$addFields": {"pct": {"$toDouble": {"$ifNull": ["$education_12th.percentage", "0"]}}}},
        {"$bucket": {"groupBy": "$pct", "boundaries": [0, 40, 50, 60, 70, 80, 90, 100, 200], "default": "Other", "output": {"count": {"$sum": 1}}}},
    ]
    try:
        score_dist = await db.student_profiles.aggregate(score_pipeline).to_list(20)
    except Exception:
        score_dist = []
    score_labels = {0: "0-40%", 40: "40-50%", 50: "50-60%", 60: "60-70%", 70: "70-80%", 80: "80-90%", 90: "90-100%", 100: "100%+"}

    # Revenue by branch
    student_ids_q = match.copy()
    student_ids = [s["id"] async for s in db.student_profiles.find(student_ids_q, {"id": 1})]
    rev_pipeline = [
        {"$match": {"student_id": {"$in": student_ids}} if student_ids else {}},
        {"$lookup": {"from": "student_profiles", "localField": "student_id", "foreignField": "id", "as": "student"}},
        {"$unwind": {"path": "$student", "preserveNullAndEmptyArrays": True}},
        {"$group": {"_id": "$student.branch_id", "total": {"$sum": "$amount"}}}
    ]
    rev_dist = await db.payments.aggregate(rev_pipeline).to_list(100)
    revenue_by_branch = []
    for r in rev_dist:
        if r["_id"]:
            br = await db.branches.find_one({"id": r["_id"]}, {"_id": 0, "name": 1})
            revenue_by_branch.append({"branch": br.get("name", "Unknown") if br else "Unknown", "revenue": r["total"]})

    trends = await db.student_profiles.aggregate(base + [
        {"$addFields": {"date": {"$substr": ["$created_at", 0, 10]}}},
        {"$group": {"_id": "$date", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}}, {"$limit": 30}
    ]).to_list(30)

    return {
        "counselling_type_distribution": [{"name": t["_id"] or "Unknown", "value": t["count"]} for t in type_dist],
        "verification_status_distribution": [{"name": s["_id"] or "PENDING", "value": s["count"]} for s in status_dist],
        "category_distribution": [{"name": c["_id"] or "Unknown", "value": c["count"]} for c in cat_dist],
        "location_distribution": [{"name": l["_id"] or "Unknown", "value": l["count"]} for l in location_dist if l["_id"]],
        "score_distribution": [{"name": score_labels.get(s["_id"], str(s["_id"])), "value": s["count"]} for s in score_dist if s["_id"] != "Other"],
        "revenue_by_branch": revenue_by_branch,
        "registration_trends": [{"date": t["_id"], "count": t["count"]} for t in trends]
    }

@api_router.get("/admin/students-drilldown")
async def admin_drilldown(
    date_from: str = Query(""), date_to: str = Query(""),
    branch_id: str = Query(""), counselling_type: str = Query(""),
    stream: str = Query(""),
    category: str = Query(""), verification_status: str = Query(""),
    page: int = Query(1, ge=1), limit: int = Query(50, ge=1, le=200),
    user: dict = Depends(require_role("ADMIN"))
):
    query = {}
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    if branch_id:
        query["branch_id"] = branch_id
    if counselling_type:
        query["counselling_type"] = counselling_type
    if stream:
        if counselling_type == "MEDICAL":
            query["selection_display"] = {"$regex": stream, "$options": "i"}
        elif counselling_type == "ENGINEERING":
            query["engineering_type"] = stream
    if category:
        query["category"] = category
    if verification_status:
        query["verification_status"] = verification_status
    skip = (page - 1) * limit
    total = await db.student_profiles.count_documents(query)
    students = await db.student_profiles.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    # Enrich with branch names
    for s in students:
        br = await db.branches.find_one({"id": s.get("branch_id", "")}, {"_id": 0, "name": 1})
        s["branch_name"] = br.get("name", "") if br else ""
    return {"students": students, "total": total, "page": page, "limit": limit}

@api_router.get("/admin/export-students")
async def export_students(
    date_from: str = Query(""), date_to: str = Query(""),
    branch_id: str = Query(""), counselling_type: str = Query(""),
    user: dict = Depends(require_role("ADMIN"))
):
    query = {}
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    if branch_id:
        query["branch_id"] = branch_id
    if counselling_type:
        query["counselling_type"] = counselling_type
    students = await db.student_profiles.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    headers = ["Reg No", "Name", "Phone", "Parent Phone", "Gender", "DOB", "District", "Category", "Caste", "Counselling Type", "Selection", "Status", "Verification", "Created"]
    ws.append(headers)
    for s in students:
        ws.append([
            s.get("registration_no", ""), s.get("full_name", ""), s.get("phone", ""),
            s.get("parent_phone", ""), s.get("gender", ""), s.get("dob", ""),
            s.get("district", ""), s.get("category", ""), s.get("caste", ""),
            s.get("counselling_type", ""), s.get("selection_display", ""),
            s.get("status", ""), s.get("verification_status", ""), s.get("created_at", "")[:10]
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await log_audit(user["id"], "", "EXPORT_DATA", "", f"Exported {len(students)} student records")
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           headers={"Content-Disposition": "attachment; filename=students_export.xlsx"})

@api_router.get("/admin/audit-logs")
async def admin_audit_logs(page: int = Query(1, ge=1), limit: int = Query(50, ge=1, le=100), user: dict = Depends(require_role("ADMIN"))):
    skip = (page - 1) * limit
    total = await db.audit_logs.count_documents({})
    logs = await db.audit_logs.find({}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"logs": logs, "total": total, "page": page, "limit": limit}

# ==================== BRANCH ROUTES ====================

@api_router.get("/branch/dashboard")
async def branch_dashboard(user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    total = await db.student_profiles.count_documents({"branch_id": bid})
    verified = await db.student_profiles.count_documents({"branch_id": bid, "verification_status": "VERIFIED"})
    pending = await db.student_profiles.count_documents({"branch_id": bid, "verification_status": {"$in": ["PENDING", "CORRECTION_REQUIRED"]}})
    rev_pipeline = [
        {"$lookup": {"from": "student_profiles", "localField": "student_id", "foreignField": "id", "as": "student"}},
        {"$unwind": "$student"}, {"$match": {"student.branch_id": bid}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}, "pending_total": {"$sum": "$pending_amount"}}}
    ]
    rev_result = await db.payments.aggregate(rev_pipeline).to_list(1)
    revenue = rev_result[0]["total"] if rev_result else 0
    pending_amount = rev_result[0].get("pending_total", 0) if rev_result else 0

    exp_result = await db.branch_expenses.aggregate([
        {"$match": {"branch_id": bid, "is_deleted": False}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    total_expense = exp_result[0]["total"] if exp_result else 0

    recent = await db.student_profiles.find({"branch_id": bid}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    
    return {
        "total_students": total, "verified_students": verified, "pending_verification": pending,
        "total_revenue": revenue, "total_pending_amount": pending_amount,
        "total_expense": total_expense, "recent_students": recent
    }

@api_router.post("/branch/students/register")
async def register_student(data: StudentRegistration, user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    existing = await db.student_profiles.find_one({"phone": data.phone, "branch_id": bid})
    if existing:
        raise HTTPException(status_code=400, detail="Student with this phone already registered")
    reg_no = await generate_reg_no(data.counselling_type, bid)
    student_email = f"{data.phone}@ameportal.in"
    existing_user = await db.users.find_one({"email": student_email})
    if existing_user:
        student_email = f"{data.phone}.{uuid.uuid4().hex[:4]}@ameportal.in"
    student_user_id = str(uuid.uuid4())
    student_user = {
        "id": student_user_id, "email": student_email,
        "password": hash_password(data.phone), "name": data.full_name,
        "role": "STUDENT", "branch_id": bid, "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(student_user)
    student_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    profile = {
        "id": student_id, "user_id": student_user_id, "branch_id": bid,
        "registration_no": reg_no, "full_name": data.full_name,
        "phone": data.phone, "parent_phone": data.parent_phone,
        "gender": data.gender, "dob": data.dob, "district": data.district,
        "category": data.category, "caste": data.caste,
        "special_reservation": data.special_reservation,
        "fresher_repeater": data.fresher_repeater,
        "counselling_type": data.counselling_type,
        "medical_selections": data.medical_selections,
        "selection_display": data.selection_display,
        "engineering_type": data.engineering_type,
        "branch_priority": data.branch_priority,
        "registered_by": data.registered_by or user.get("name", ""),
        "registered_by_id": data.registered_by_id or user["id"],
        "status": "REGISTERED", "verification_status": "PENDING",
        "additional_details": {}, "education_10th": {}, "education_11th": {},
        "education_12th": {}, "parent_details": {}, "address_details": {},
        "has_additional_details": False,
        "created_at": now, "updated_at": now
    }
    await db.student_profiles.insert_one(profile)
    if data.payment_amount > 0 or data.pending_amount > 0:
        await db.payments.insert_one({
            "id": str(uuid.uuid4()), "student_id": student_id,
            "amount": data.payment_amount, "pending_amount": data.pending_amount,
            "concession_amount": data.concession_amount,
            "payment_mode": data.payment_mode,
            "utr_number": data.utr_number, "note": data.payment_note,
            "created_at": now
        })
    doc_types = DOCUMENT_TYPES.get(data.counselling_type, [])
    for dt in doc_types:
        await db.documents.insert_one({
            "id": str(uuid.uuid4()), "student_id": student_id,
            "document_type": dt, "file_url": "", "status": "PENDING",
            "remarks": "", "created_at": now, "updated_at": now
        })
    # Create v1 Snapshot
    snapshot_data = {
        "full_name": profile["full_name"], "phone": profile["phone"], "parent_phone": profile["parent_phone"],
        "gender": profile["gender"], "dob": profile["dob"], "district": profile["district"],
        "category": profile["category"], "caste": profile["caste"],
        "special_reservation": profile["special_reservation"], "fresher_repeater": profile["fresher_repeater"],
        "counselling_type": profile["counselling_type"], "medical_selections": profile["medical_selections"],
        "selection_display": profile["selection_display"], "engineering_type": profile["engineering_type"],
        "branch_priority": profile["branch_priority"], "registered_by": profile["registered_by"],
        "registered_by_id": profile["registered_by_id"], "payment_amount": data.payment_amount,
        "pending_amount": data.pending_amount, "concession_amount": data.concession_amount,
        "payment_mode": data.payment_mode, "utr_number": data.utr_number, "payment_note": data.payment_note
    }
    await db.registration_history.insert_one({
        "id": str(uuid.uuid4()), "student_id": student_id, "version": 1,
        "snapshot": snapshot_data, "edited_by": user["id"], "created_at": now
    })

    await log_audit(user["id"], bid, "REGISTER_STUDENT", student_id, f"Registered: {data.full_name} ({reg_no})")
    profile.pop("_id", None)
    return {"student": profile, "credentials": {"email": student_email, "password": data.phone}}

@api_router.post("/branch/students/bulk-import")
async def bulk_import_students(file: UploadFile = File(...), user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    contents = await file.read()
    decoded = contents.decode("utf-8")
    io_string = io.StringIO(decoded)
    reader = csv.DictReader(io_string)
    
    success_count = 0
    failed_count = 0
    errors = []
    
    for row_idx, row in enumerate(reader, start=2):
        try:
            phone = row.get("phone", "").strip()
            if not phone:
                failed_count += 1
                errors.append(f"Row {row_idx}: Mobile number is required")
                continue
            
            # Check existing
            existing = await db.student_profiles.find_one({"phone": phone, "branch_id": bid})
            if existing:
                failed_count += 1
                errors.append(f"Row {row_idx}: Student with phone {phone} already registered")
                continue

            # Parse DOB (DD/MM/YYYY to YYYY-MM-DD)
            dob = row.get("dob", "").strip()
            if "/" in dob:
                parts = dob.split("/")
                if len(parts) == 3:
                     # Assume DD/MM/YYYY, convert to YYYY-MM-DD
                     dob = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"

            # Parse Medical Selections (Grouped for Frontend)
            medical_selections = []
            selection_display = ""
            med_str = row.get("medical_selections", "").strip()
            c_type = row.get("counselling_type", "MEDICAL").upper().strip()
            
            if c_type == "MEDICAL" and med_str:
                temp_map = {} # type -> {courses: [], states: []}
                selections = med_str.split(",")
                for s in selections:
                    parts = s.split("|")
                    if len(parts) >= 2:
                        cat = parts[0].strip()
                        course = parts[1].strip()
                        state = parts[2].strip() if len(parts) > 2 else ""
                        
                        if cat not in temp_map:
                            temp_map[cat] = {"type": cat, "courses": [], "states": []}
                        
                        if course and course not in temp_map[cat]["courses"]:
                            temp_map[cat]["courses"].append(course)
                        if state and state not in temp_map[cat]["states"]:
                            temp_map[cat]["states"].append(state)
                
                medical_selections = list(temp_map.values())
                
                # Generate display
                display_parts = []
                for m in medical_selections:
                    display_parts.append(f"{m['type']} ({', '.join(m['courses'])})")
                selection_display = f"Medical - {', '.join(display_parts)}"
            
            elif c_type == "ENGINEERING":
                selection_display = f"Engg - {row.get('engineering_type','')}"
            elif c_type == "DSE":
                selection_display = "DSE"

            # Prepare StudentRegistration data
            data = StudentRegistration(
                full_name=row.get("full_name", "").strip(),
                phone=phone,
                parent_phone=row.get("parent_phone", "").strip(),
                gender=row.get("gender", "Male").strip(),
                dob=dob,
                district=row.get("district", "").strip(),
                category=row.get("category", "GEN").strip(),
                caste=row.get("caste", "").strip(),
                special_reservation=row.get("special_reservation", "None").strip(),
                fresher_repeater=row.get("fresher_repeater", "Fresher").strip(),
                counselling_type=c_type,
                medical_selections=medical_selections,
                selection_display=selection_display,
                engineering_type=row.get("engineering_type", "").strip(),
                branch_priority=row.get("branch_priority", "").strip(),
                payment_amount=float(row.get("payment_amount", 0) or 0),
                pending_amount=float(row.get("pending_amount", 0) or 0),
                concession_amount=float(row.get("concession_amount", 0) or 0),
                payment_mode=row.get("payment_mode", "CASH").strip(),
                utr_number=row.get("utr_number", "").strip(),
                payment_note=row.get("payment_note", "").strip()
            )

            # --- COPY OF REGISTRATION LOGIC ---
            reg_no = await generate_reg_no(data.counselling_type, bid)
            student_email = f"{data.phone}@ameportal.in"
            existing_user = await db.users.find_one({"email": student_email})
            if existing_user:
                student_email = f"{data.phone}.{uuid.uuid4().hex[:4]}@ameportal.in"
            
            student_user_id = str(uuid.uuid4())
            student_user = {
                "id": student_user_id, "email": student_email,
                "password": hash_password(data.phone), "name": data.full_name,
                "role": "STUDENT", "branch_id": bid, "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(student_user)
            student_id = str(uuid.uuid4())
            now = datetime.now(timezone.utc).isoformat()
            profile = {
                "id": student_id, "user_id": student_user_id, "branch_id": bid,
                "registration_no": reg_no, "full_name": data.full_name,
                "phone": data.phone, "parent_phone": data.parent_phone,
                "gender": data.gender, "dob": data.dob, "district": data.district,
                "category": data.category, "caste": data.caste,
                "special_reservation": data.special_reservation,
                "fresher_repeater": data.fresher_repeater,
                "counselling_type": data.counselling_type,
                "medical_selections": data.medical_selections,
                "selection_display": data.selection_display,
                "engineering_type": data.engineering_type,
                "branch_priority": data.branch_priority,
                "registered_by": user.get("name", ""),
                "registered_by_id": user["id"],
                "status": "REGISTERED", "verification_status": "PENDING",
                "additional_details": {}, "education_10th": {}, "education_11th": {},
                "education_12th": {}, "parent_details": {}, "address_details": {},
                "has_additional_details": False,
                "created_at": now, "updated_at": now
            }
            await db.student_profiles.insert_one(profile)
            if data.payment_amount > 0 or data.pending_amount > 0:
                await db.payments.insert_one({
                    "id": str(uuid.uuid4()), "student_id": student_id,
                    "amount": data.payment_amount, "pending_amount": data.pending_amount,
                    "concession_amount": data.concession_amount,
                    "payment_mode": data.payment_mode,
                    "utr_number": data.utr_number, "note": data.payment_note,
                    "created_at": now
                })
            doc_types = DOCUMENT_TYPES.get(data.counselling_type, [])
            for dt in doc_types:
                await db.documents.insert_one({
                    "id": str(uuid.uuid4()), "student_id": student_id,
                    "document_type": dt, "file_url": "", "status": "PENDING",
                    "remarks": "", "created_at": now, "updated_at": now
                })
            
            # History Snapshot
            snapshot_data = {
                "full_name": profile["full_name"], "phone": profile["phone"], "parent_phone": profile["parent_phone"],
                "gender": profile["gender"], "dob": profile["dob"], "district": profile["district"],
                "category": profile["category"], "caste": profile["caste"],
                "special_reservation": profile["special_reservation"], "fresher_repeater": profile["fresher_repeater"],
                "counselling_type": profile["counselling_type"], "medical_selections": profile["medical_selections"],
                "selection_display": profile["selection_display"], "engineering_type": profile["engineering_type"],
                "branch_priority": profile["branch_priority"], "registered_by": profile["registered_by"],
                "registered_by_id": profile["registered_by_id"], "payment_amount": data.payment_amount,
                "pending_amount": data.pending_amount, "concession_amount": data.concession_amount,
                "payment_mode": data.payment_mode, "utr_number": data.utr_number, "payment_note": data.payment_note
            }
            await db.registration_history.insert_one({
                "id": str(uuid.uuid4()), "student_id": student_id, "version": 1,
                "snapshot": snapshot_data, "edited_by": user["id"], "created_at": now
            })
            
            await log_audit(user["id"], bid, "BULK_IMPORT_STUDENT", student_id, f"Imported: {data.full_name} ({reg_no})")
            success_count += 1
            
        except Exception as e:
            failed_count += 1
            errors.append(f"Row {row_idx}: Internal Error - {str(e)}")

    return {
        "success_count": success_count,
        "failed_count": failed_count,
        "errors": errors
    }

@api_router.get("/branch/students")
async def branch_students(
    search: str = Query(""), status: str = Query(""),
    counselling_type: str = Query(""), stream: str = Query(""), page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))
):
    bid = user.get("branch_id")
    query = {"branch_id": bid}
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"registration_no": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}}
        ]
    if status:
        query["verification_status"] = status
    if counselling_type:
        query["counselling_type"] = counselling_type
    if stream:
        if counselling_type == "MEDICAL":
            query["selection_display"] = {"$regex": stream, "$options": "i"}
        elif counselling_type == "ENGINEERING":
            query["engineering_type"] = stream
    skip = (page - 1) * limit
    total = await db.student_profiles.count_documents(query)
    students = await db.student_profiles.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"students": students, "total": total, "page": page, "limit": limit}

@api_router.get("/branch/students/{student_id}")
async def get_branch_student(student_id: str, user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    student = await db.student_profiles.find_one({"id": student_id, "branch_id": bid}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    documents = await db.documents.find({"student_id": student_id}, {"_id": 0}).to_list(50)
    payment = await db.payments.find_one({"student_id": student_id}, {"_id": 0})
    vlogs = await db.verification_logs.find({"student_id": student_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"student": student, "documents": documents, "payment": payment, "verification_logs": vlogs}

@api_router.get("/branch/verification-queue")
async def verification_queue(user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    students = await db.student_profiles.find(
        {"branch_id": bid, "$or": [
            {"status": "DOCS_UPLOADED", "verification_status": "PENDING"},
            {"verification_status": "CORRECTION_REQUIRED"}
        ]}, {"_id": 0}
    ).sort("updated_at", 1).to_list(100)
    return students

@api_router.get("/branch/verification/{student_id}")
async def get_verification(student_id: str, user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    student = await db.student_profiles.find_one({"id": student_id, "branch_id": bid}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    documents = await db.documents.find({"student_id": student_id}, {"_id": 0}).to_list(50)
    prev_logs = await db.verification_logs.find({"student_id": student_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"student": student, "documents": documents, "previous_logs": prev_logs}

@api_router.post("/branch/verification/{student_id}")
async def submit_verification(student_id: str, data: VerificationSubmit, user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    student = await db.student_profiles.find_one({"id": student_id, "branch_id": bid})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    all_correct = True
    now = datetime.now(timezone.utc).isoformat()
    for d in data.decisions:
        if d.decision == "INCORRECT":
            all_correct = False
            if not d.comment:
                raise HTTPException(status_code=400, detail=f"Comment required for: {d.field_name}")
        await db.verification_logs.insert_one({
            "id": str(uuid.uuid4()), "student_id": student_id,
            "verified_by": user["id"], "verifier_name": user.get("name", ""),
            "field_name": d.field_name, "decision": d.decision,
            "comment": d.comment, "created_at": now
        })
    new_status = "VERIFIED" if all_correct else "CORRECTION_REQUIRED"
    update = {"verification_status": new_status, "updated_at": now}
    if all_correct:
        update["status"] = "VERIFIED"
    await db.student_profiles.update_one({"id": student_id}, {"$set": update})
    for d in data.decisions:
        if d.field_name.startswith("doc_"):
            doc_type = d.field_name.replace("doc_", "").replace("_", " ").title()
            doc_status = "VERIFIED" if d.decision == "CORRECT" else "CORRECTION_REQUIRED"
            await db.documents.update_many(
                {"student_id": student_id, "document_type": {"$regex": doc_type, "$options": "i"}},
                {"$set": {"status": doc_status, "remarks": d.comment, "updated_at": now}}
            )
    await log_audit(user["id"], bid, "VERIFY_STUDENT", student_id, f"Verification: {new_status} for {student.get('full_name', '')}")
    return {"status": new_status, "message": f"Verification: {new_status}"}

@api_router.get("/branch/audit-logs")
async def branch_audit_logs(page: int = Query(1, ge=1), limit: int = Query(50, ge=1, le=100), user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    skip = (page - 1) * limit
    total = await db.audit_logs.count_documents({"branch_id": bid})
    logs = await db.audit_logs.find({"branch_id": bid}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"logs": logs, "total": total, "page": page, "limit": limit}

# ==================== BRANCH STAFF MANAGEMENT ====================

@api_router.get("/branch/staff")
async def list_staff(user: dict = Depends(require_role("BRANCH_HEAD"))):
    bid = user.get("branch_id")
    staff = await db.users.find({"branch_id": bid, "role": "STAFF"}, {"_id": 0, "password": 0}).sort("created_at", -1).to_list(100)
    return staff

@api_router.post("/branch/staff")
async def create_staff(data: StaffCreate, user: dict = Depends(require_role("BRANCH_HEAD"))):
    bid = user.get("branch_id")
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already in use")
    staff_user = {
        "id": str(uuid.uuid4()), "email": data.email,
        "password": hash_password(data.password), "name": data.name,
        "role": "STAFF", "branch_id": bid, "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(staff_user)
    await log_audit(user["id"], bid, "CREATE_STAFF", staff_user["id"], f"Created staff: {data.name}")
    return {"id": staff_user["id"], "email": data.email, "name": data.name, "role": "STAFF", "is_active": True}

@api_router.put("/branch/staff/{staff_id}")
async def update_staff(staff_id: str, is_active: bool = Query(...), user: dict = Depends(require_role("BRANCH_HEAD"))):
    bid = user.get("branch_id")
    result = await db.users.update_one({"id": staff_id, "branch_id": bid, "role": "STAFF"}, {"$set": {"is_active": is_active}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Staff not found")
    await log_audit(user["id"], bid, "UPDATE_STAFF", staff_id, f"Staff {'activated' if is_active else 'deactivated'}")
    return {"message": "Staff updated"}

@api_router.delete("/branch/staff/{staff_id}")
async def delete_staff(staff_id: str, user: dict = Depends(require_role("BRANCH_HEAD"))):
    bid = user.get("branch_id")
    result = await db.users.delete_one({"id": staff_id, "branch_id": bid, "role": "STAFF"})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Staff not found")
    await log_audit(user["id"], bid, "DELETE_STAFF", staff_id, "Deleted staff account")
    return {"message": "Staff deleted"}

@api_router.get("/branch/staff-list")
async def branch_staff_list(user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    staff = await db.users.find({"branch_id": bid, "role": {"$in": ["BRANCH_HEAD", "STAFF"]}, "is_active": True}, {"_id": 0, "id": 1, "name": 1, "role": 1}).to_list(50)
    return staff

# ==================== REGISTRATION WORKSPACE & VERSIONING ====================

@api_router.get("/branch/students/{student_id}/history")
async def get_registration_history(student_id: str, user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    student = await db.student_profiles.find_one({"id": student_id, "branch_id": bid})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    history = await db.registration_history.find({"student_id": student_id}, {"_id": 0}).sort("version", -1).to_list(100)
    return history

@api_router.put("/branch/students/{student_id}/registration")
async def update_student_registration(student_id: str, data: RegistrationUpdate, user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    profile = await db.student_profiles.find_one({"id": student_id, "branch_id": bid})
    if not profile:
        raise HTTPException(status_code=404, detail="Student not found")

    last_version = await db.registration_history.find_one({"student_id": student_id}, sort=[("version", -1)])
    next_version = (last_version["version"] + 1) if last_version else 1

    now = datetime.now(timezone.utc).isoformat()
    update_data = {k: v for k, v in data.model_dump().items() if v is not None and k not in ["updated_by_id", "edit_reason"]}
    
    # Separate payment from profile
    payment_fields = ["payment_amount", "pending_amount", "concession_amount", "payment_mode", "utr_number", "payment_note"]
    profile_update = {k: v for k, v in update_data.items() if k not in payment_fields}
    payment_update = {k: v for k, v in update_data.items() if k in payment_fields}

    if profile_update:
        profile_update["updated_at"] = now
        await db.student_profiles.update_one({"id": student_id}, {"$set": profile_update})

    # Update or Create Payment record
    if payment_update:
        await db.payments.update_one(
            {"student_id": student_id},
            {"$set": {
                "amount": payment_update.get("payment_amount", 0),
                "pending_amount": payment_update.get("pending_amount", 0),
                "concession_amount": payment_update.get("concession_amount", 0),
                "payment_mode": payment_update.get("payment_mode", "CASH"),
                "utr_number": payment_update.get("utr_number", ""),
                "note": payment_update.get("payment_note", ""),
                "updated_at": now
            }},
            upsert=True
        )

    # Re-fetch latest to create complete snapshot
    latest_profile = await db.student_profiles.find_one({"id": student_id}, {"_id": 0})
    latest_payment = await db.payments.find_one({"student_id": student_id}, {"_id": 0})
    
    snapshot = {
        "full_name": latest_profile.get("full_name"), "phone": latest_profile.get("phone"),
        "parent_phone": latest_profile.get("parent_phone"), "gender": latest_profile.get("gender"),
        "dob": latest_profile.get("dob"), "district": latest_profile.get("district"),
        "category": latest_profile.get("category"), "caste": latest_profile.get("caste"),
        "special_reservation": latest_profile.get("special_reservation"),
        "fresher_repeater": latest_profile.get("fresher_repeater"),
        "counselling_type": latest_profile.get("counselling_type"),
        "medical_selections": latest_profile.get("medical_selections"),
        "selection_display": latest_profile.get("selection_display"),
        "engineering_type": latest_profile.get("engineering_type"),
        "branch_priority": latest_profile.get("branch_priority"),
        "registered_by": latest_profile.get("registered_by"),
        "registered_by_id": latest_profile.get("registered_by_id"),
        "payment_amount": latest_payment.get("amount", 0) if latest_payment else 0,
        "payment_mode": latest_payment.get("payment_mode", "CASH") if latest_payment else "CASH",
        "utr_number": latest_payment.get("utr_number", "") if latest_payment else "",
        "payment_note": latest_payment.get("note", "") if latest_payment else "",
        "updated_by_id": data.updated_by_id or user["id"],
        "edit_reason": data.edit_reason or ""
    }

    await db.registration_history.insert_one({
        "id": str(uuid.uuid4()), "student_id": student_id, "version": next_version,
        "snapshot": snapshot, "edited_by": user["id"], "created_at": now
    })

    audit_note = f"Updated registration details to v{next_version}. Reason: {data.edit_reason}" if data.edit_reason else f"Updated registration details to v{next_version}"
    await log_audit(user["id"], bid, "UPDATE_REGISTRATION", student_id, audit_note)
    return {"message": f"Successfully updated to v{next_version}", "version": next_version}

# ==================== BRANCH ANALYTICS ====================

@api_router.get("/branch/analytics")
async def branch_analytics(
    date_from: str = Query(""), date_to: str = Query(""),
    counselling_type: str = Query(""),
    user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))
):
    bid = user.get("branch_id")
    match = {"branch_id": bid}
    if date_from:
        match["created_at"] = {"$gte": date_from}
    if date_to:
        match.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    if counselling_type:
        match["counselling_type"] = counselling_type
    base = [{"$match": match}]

    type_dist = await db.student_profiles.aggregate(base + [{"$group": {"_id": "$counselling_type", "count": {"$sum": 1}}}]).to_list(10)
    status_dist = await db.student_profiles.aggregate(base + [{"$group": {"_id": "$verification_status", "count": {"$sum": 1}}}]).to_list(10)
    cat_dist = await db.student_profiles.aggregate(base + [{"$group": {"_id": "$category", "count": {"$sum": 1}}}]).to_list(20)
    location_dist = await db.student_profiles.aggregate(base + [{"$group": {"_id": "$district", "count": {"$sum": 1}}}, {"$sort": {"count": -1}}, {"$limit": 20}]).to_list(20)

    score_pipeline = base + [
        {"$match": {"education_12th.percentage": {"$exists": True, "$ne": ""}}},
        {"$addFields": {"pct": {"$toDouble": {"$ifNull": ["$education_12th.percentage", "0"]}}}},
        {"$bucket": {"groupBy": "$pct", "boundaries": [0, 40, 50, 60, 70, 80, 90, 100, 200], "default": "Other", "output": {"count": {"$sum": 1}}}},
    ]
    try:
        score_dist = await db.student_profiles.aggregate(score_pipeline).to_list(20)
    except Exception:
        score_dist = []
    score_labels = {0: "0-40%", 40: "40-50%", 50: "50-60%", 60: "60-70%", 70: "70-80%", 80: "80-90%", 90: "90-100%"}

    trends = await db.student_profiles.aggregate(base + [
        {"$addFields": {"date": {"$substr": ["$created_at", 0, 10]}}},
        {"$group": {"_id": "$date", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}}, {"$limit": 30}
    ]).to_list(30)

    # Revenue
    student_ids = [s["id"] async for s in db.student_profiles.find(match, {"id": 1})]
    rev_result = await db.payments.aggregate([
        {"$match": {"student_id": {"$in": student_ids}} if student_ids else {"student_id": "__none__"}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    revenue = rev_result[0]["total"] if rev_result else 0

    return {
        "total_students": await db.student_profiles.count_documents(match),
        "verified_students": await db.student_profiles.count_documents({**match, "verification_status": "VERIFIED"}),
        "pending_students": await db.student_profiles.count_documents({**match, "verification_status": {"$in": ["PENDING", "CORRECTION_REQUIRED"]}}),
        "total_revenue": revenue,
        "counselling_type_distribution": [{"name": t["_id"] or "Unknown", "value": t["count"]} for t in type_dist],
        "verification_status_distribution": [{"name": s["_id"] or "PENDING", "value": s["count"]} for s in status_dist],
        "category_distribution": [{"name": c["_id"] or "Unknown", "value": c["count"]} for c in cat_dist],
        "location_distribution": [{"name": l["_id"] or "Unknown", "value": l["count"]} for l in location_dist if l["_id"]],
        "score_distribution": [{"name": score_labels.get(s["_id"], str(s["_id"])), "value": s["count"]} for s in score_dist if s["_id"] != "Other"],
        "registration_trends": [{"date": t["_id"], "count": t["count"]} for t in trends]
    }

@api_router.get("/branch/export-students")
async def export_branch_students(
    date_from: str = Query(""), date_to: str = Query(""),
    counselling_type: str = Query(""),
    user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))
):
    bid = user.get("branch_id")
    query = {"branch_id": bid}
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    if counselling_type:
        query["counselling_type"] = counselling_type
    students = await db.student_profiles.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    headers = ["Reg No", "Name", "Phone", "Parent Phone", "Gender", "DOB", "District", "Category", "Caste", "Counselling Type", "Selection", "Status", "Verification", "Created"]
    ws.append(headers)
    for s in students:
        ws.append([s.get("registration_no", ""), s.get("full_name", ""), s.get("phone", ""), s.get("parent_phone", ""), s.get("gender", ""), s.get("dob", ""), s.get("district", ""), s.get("category", ""), s.get("caste", ""), s.get("counselling_type", ""), s.get("selection_display", ""), s.get("status", ""), s.get("verification_status", ""), s.get("created_at", "")[:10]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await log_audit(user["id"], bid, "EXPORT_DATA", "", f"Exported {len(students)} student records")
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=branch_students_export.xlsx"})

# ==================== STUDENT ADDITIONAL DETAILS ====================

@api_router.put("/branch/students/{student_id}/additional")
async def update_student_additional(student_id: str, data: StudentAdditionalUpdate, user: dict = Depends(require_role("BRANCH_HEAD", "STAFF", "STUDENT"))):
    if user["role"] == "STUDENT":
        profile = await db.student_profiles.find_one({"user_id": user["id"]}, {"_id": 0, "id": 1})
        if not profile or profile["id"] != student_id:
            raise HTTPException(status_code=403, detail="Access denied")
    else:
        bid = user.get("branch_id")
        profile = await db.student_profiles.find_one({"id": student_id, "branch_id": bid})
        if not profile:
            raise HTTPException(status_code=404, detail="Student not found")

    now = datetime.now(timezone.utc).isoformat()
    update = {
        "additional_details": {
            "mother_name": data.mother_name, "email": data.email, "religion": data.religion,
            "family_income": data.family_income, "aadhaar_number": data.aadhaar_number,
            "address": data.address
        },
        "address_details": {"taluka": data.taluka, "district": data.address_district, "pin_code": data.pin_code},
        "education_10th": {
            "year": data.edu_10_year, "place": data.edu_10_place, "school_type": data.edu_10_school_type,
            "board": data.edu_10_board, "state": data.edu_10_state, "district": data.edu_10_district,
            "roll": data.edu_10_roll, "total_marks": data.edu_10_total, "obtained_marks": data.edu_10_obtained,
            "percentage": data.edu_10_percentage, "school": data.edu_10_school,
            "school_address": data.edu_10_school_address, "pin_code": data.edu_10_pin
        },
        "education_11th": {
            "year": data.edu_11_year, "place": data.edu_11_place, "college_type": data.edu_11_college_type,
            "board": data.edu_11_board, "state": data.edu_11_state, "district": data.edu_11_district,
            "roll": data.edu_11_roll, "total_marks": data.edu_11_total, "obtained_marks": data.edu_11_obtained,
            "percentage": data.edu_11_percentage, "college": data.edu_11_college,
            "college_address": data.edu_11_college_address, "pin_code": data.edu_11_pin
        },
        "education_12th": {
            "year": data.edu_12_year, "place": data.edu_12_place, "college_type": data.edu_12_college_type,
            "board": data.edu_12_board, "state": data.edu_12_state, "district": data.edu_12_district,
            "roll": data.edu_12_roll, "total_marks": data.edu_12_total, "obtained_marks": data.edu_12_obtained,
            "percentage": data.edu_12_percentage, "college": data.edu_12_college,
            "college_address": data.edu_12_college_address, "pin_code": data.edu_12_pin
        },
        "parent_details": {
            "father_occupation": data.father_occupation, "father_qualification": data.father_qualification,
            "mother_occupation": data.mother_occupation, "mother_qualification": data.mother_qualification
        },
        "has_additional_details": True,
        "updated_at": now
    }
    await db.student_profiles.update_one({"id": student_id}, {"$set": update})
    await log_audit(user["id"], user.get("branch_id", ""), "UPDATE_STUDENT_DETAILS", student_id, f"Updated additional details")
    updated = await db.student_profiles.find_one({"id": student_id}, {"_id": 0})
    return updated

@api_router.get("/branch/students/{student_id}/receipt")
async def get_receipt_data(student_id: str, user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    student = await db.student_profiles.find_one({"id": student_id, "branch_id": bid}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    payment = await db.payments.find_one({"student_id": student_id}, {"_id": 0})
    branch = await db.branches.find_one({"id": bid}, {"_id": 0, "name": 1, "location": 1, "contact_number": 1})
    return {"student": student, "payment": payment, "branch": branch}

# ==================== CHAT SYSTEM ====================

@api_router.get("/chat/{student_id}/messages")
async def get_chat_messages(student_id: str, user: dict = Depends(get_current_user)):
    if user["role"] == "STUDENT":
        profile = await db.student_profiles.find_one({"user_id": user["id"]}, {"_id": 0, "id": 1})
        if not profile or profile["id"] != student_id:
            raise HTTPException(status_code=403, detail="Access denied")
    elif user["role"] in ["BRANCH_HEAD", "STAFF"]:
        bid = user.get("branch_id")
        student = await db.student_profiles.find_one({"id": student_id, "branch_id": bid})
        if not student:
            raise HTTPException(status_code=404, detail="Student not found")
    else:
        raise HTTPException(status_code=403, detail="Access denied")
    messages = await db.chat_messages.find({"student_id": student_id}, {"_id": 0}).sort("created_at", 1).to_list(500)
    # Mark messages as read
    await db.chat_messages.update_many(
        {"student_id": student_id, "sender_id": {"$ne": user["id"]}, "read": False},
        {"$set": {"read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    return messages

@api_router.post("/chat/{student_id}/messages")
async def send_chat_message(student_id: str, data: ChatMessageCreate, user: dict = Depends(get_current_user)):
    if user["role"] == "STUDENT":
        profile = await db.student_profiles.find_one({"user_id": user["id"]}, {"_id": 0, "id": 1})
        if not profile or profile["id"] != student_id:
            raise HTTPException(status_code=403, detail="Access denied")
    elif user["role"] in ["BRANCH_HEAD", "STAFF"]:
        bid = user.get("branch_id")
        student = await db.student_profiles.find_one({"id": student_id, "branch_id": bid})
        if not student:
            raise HTTPException(status_code=404, detail="Student not found")
    else:
        raise HTTPException(status_code=403, detail="Access denied")
    if not data.content and not data.attachment_url:
        raise HTTPException(status_code=400, detail="Message or attachment required")
    now = datetime.now(timezone.utc).isoformat()
    msg = {
        "id": str(uuid.uuid4()), "student_id": student_id,
        "sender_id": user["id"], "sender_name": user.get("name", ""),
        "sender_role": user["role"], "content": data.content,
        "attachment_url": data.attachment_url, "read": False,
        "created_at": now
    }
    await db.chat_messages.insert_one(msg)
    # Create notification for the other party
    if user["role"] == "STUDENT":
        # Notify branch staff
        branch_users = await db.users.find({"branch_id": user.get("branch_id"), "role": {"$in": ["BRANCH_HEAD", "STAFF"]}}, {"id": 1}).to_list(50)
        for bu in branch_users:
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()), "user_id": bu["id"], "student_id": student_id,
                "title": "New Message", "message": f"New message from student",
                "type": "chat", "is_read": False, "created_at": now
            })
    else:
        student_profile = await db.student_profiles.find_one({"id": student_id}, {"user_id": 1})
        if student_profile:
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()), "user_id": student_profile["user_id"], "student_id": student_id,
                "title": "New Message", "message": f"New message from staff",
                "type": "chat", "is_read": False, "created_at": now
            })
    msg.pop("_id", None)
    return msg

@api_router.post("/chat/{student_id}/upload")
async def upload_chat_attachment(student_id: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    ext = Path(file.filename).suffix if file.filename else ".pdf"
    filename = f"chat_{uuid.uuid4()}{ext}"
    filepath = UPLOAD_DIR / filename
    with open(filepath, "wb") as f:
        content = await file.read()
        f.write(content)
    return {"file_url": f"/api/files/{filename}", "filename": file.filename}

# ==================== NOTIFICATIONS ====================

@api_router.get("/notifications")
async def get_notifications(user: dict = Depends(get_current_user)):
    notifs = await db.notifications.find({"user_id": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)
    unread = await db.notifications.count_documents({"user_id": user["id"], "is_read": False})
    return {"notifications": notifs, "unread_count": unread}

@api_router.put("/notifications/{notif_id}/read")
async def mark_notification_read(notif_id: str, user: dict = Depends(get_current_user)):
    await db.notifications.update_one({"id": notif_id, "user_id": user["id"]}, {"$set": {"is_read": True}})
    return {"message": "Marked as read"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read(user: dict = Depends(get_current_user)):
    await db.notifications.update_many({"user_id": user["id"], "is_read": False}, {"$set": {"is_read": True}})
    return {"message": "All marked as read"}

# ==================== CONSTANTS API ====================

@api_router.get("/constants")
async def get_constants():
    return {
        "maharashtra_districts": MAHARASHTRA_DISTRICTS,
        "categories": CATEGORIES,
        "special_reservations": SPECIAL_RESERVATIONS,
        "medical_counselling_types": MEDICAL_COUNSELLING_TYPES,
        "other_state_mbbs_states": OTHER_STATE_MBBS_STATES,
        "other_state_bams_states": OTHER_STATE_BAMS_STATES
    }

# ==================== STUDENT ROUTES ====================

@api_router.get("/student/dashboard")
async def student_dashboard(user: dict = Depends(require_role("STUDENT"))):
    profile = await db.student_profiles.find_one({"user_id": user["id"]}, {"_id": 0})
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    documents = await db.documents.find({"student_id": profile["id"]}, {"_id": 0}).to_list(50)
    branch = await db.branches.find_one({"id": profile.get("branch_id", "")}, {"_id": 0, "name": 1, "location": 1})
    return {"profile": profile, "documents": documents, "branch": branch}

@api_router.get("/student/profile")
async def student_profile(user: dict = Depends(require_role("STUDENT"))):
    profile = await db.student_profiles.find_one({"user_id": user["id"]}, {"_id": 0})
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    return profile

@api_router.get("/student/documents")
async def student_documents(user: dict = Depends(require_role("STUDENT"))):
    profile = await db.student_profiles.find_one({"user_id": user["id"]}, {"_id": 0, "id": 1})
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    return await db.documents.find({"student_id": profile["id"]}, {"_id": 0}).to_list(50)

@api_router.get("/student/payments")
async def student_payments(user: dict = Depends(require_role("STUDENT"))):
    profile = await db.student_profiles.find_one({"user_id": user["id"]}, {"_id": 0, "id": 1})
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    return await db.payments.find({"student_id": profile["id"]}, {"_id": 0}).sort("created_at", -1).to_list(50)

@api_router.get("/student/notifications")
async def student_notifications(user: dict = Depends(require_role("STUDENT"))):
    notifs = await db.notifications.find({"user_id": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)
    unread = await db.notifications.count_documents({"user_id": user["id"], "is_read": False})
    return {"notifications": notifs, "unread_count": unread}

@api_router.post("/student/documents/upload")
async def upload_document(document_id: str = Form(...), file: UploadFile = File(...), user: dict = Depends(require_role("STUDENT"))):
    profile = await db.student_profiles.find_one({"user_id": user["id"]}, {"_id": 0, "id": 1})
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    doc = await db.documents.find_one({"id": document_id, "student_id": profile["id"]})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    ext = Path(file.filename).suffix if file.filename else ".pdf"
    filename = f"{uuid.uuid4()}{ext}"
    filepath = UPLOAD_DIR / filename
    with open(filepath, "wb") as f:
        content = await file.read()
        f.write(content)
    now = datetime.now(timezone.utc).isoformat()
    await db.documents.update_one({"id": document_id}, {"$set": {"file_url": f"/api/files/{filename}", "status": "PENDING", "remarks": "", "updated_at": now}})
    all_docs = await db.documents.find({"student_id": profile["id"]}, {"_id": 0}).to_list(50)
    all_uploaded = all(d.get("file_url") for d in all_docs)
    if all_uploaded:
        await db.student_profiles.update_one({"id": profile["id"]}, {"$set": {"status": "DOCS_UPLOADED", "updated_at": now}})
    updated_doc = await db.documents.find_one({"id": document_id}, {"_id": 0})
    return updated_doc

@api_router.get("/files/{filename}")
async def serve_file(filename: str, user: dict = Depends(get_current_user)):
    filepath = UPLOAD_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(filepath)

# ==================== EXPENSE MANAGER ====================

class ExpenseCreate(BaseModel):
    expense_for: str
    amount: float
    note: Optional[str] = ""
    payment_mode: Optional[str] = ""
    payment_done_by: str
    payment_done_by_name: str
    date: str
    attachment_url: Optional[str] = ""

class ExpenseUpdate(BaseModel):
    expense_for: Optional[str] = None
    amount: Optional[float] = None
    note: Optional[str] = None
    payment_mode: Optional[str] = None
    payment_done_by: Optional[str] = None
    payment_done_by_name: Optional[str] = None
    date: Optional[str] = None
    attachment_url: Optional[str] = None

@api_router.post("/branch/expenses")
async def create_expense(data: ExpenseCreate, user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    if data.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be positive")
    now = datetime.now(timezone.utc).isoformat()
    expense = {
        "id": str(uuid.uuid4()),
        "branch_id": bid,
        "expense_for": data.expense_for,
        "amount": data.amount,
        "note": data.note or "",
        "payment_mode": data.payment_mode or "",
        "payment_done_by": data.payment_done_by,
        "payment_done_by_name": data.payment_done_by_name,
        "date": data.date,
        "attachment_url": data.attachment_url or "",
        "created_by_id": user["id"],
        "created_by_name": user.get("name", ""),
        "last_updated_by": user["id"],
        "last_updated_at": now,
        "is_deleted": False,
        "deleted_at": None,
        "created_at": now
    }
    await db.branch_expenses.insert_one(expense)
    await log_audit(user["id"], bid, "CREATE_EXPENSE", expense["id"],
                    f"Added expense: {data.expense_for} ₹{data.amount}")
    expense.pop("_id", None)
    return expense

@api_router.get("/branch/expenses/summary")
async def expense_summary(user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    now = datetime.now(timezone.utc)
    today_str = now.strftime("%Y-%m-%d")
    month_start = now.strftime("%Y-%m-01")

    today_result = await db.branch_expenses.aggregate([
        {"$match": {"branch_id": bid, "is_deleted": False, "date": today_str}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)

    month_result = await db.branch_expenses.aggregate([
        {"$match": {"branch_id": bid, "is_deleted": False, "date": {"$gte": month_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)

    return {
        "today": today_result[0]["total"] if today_result else 0,
        "month": month_result[0]["total"] if month_result else 0
    }

@api_router.get("/branch/expenses")
async def list_expenses(
    date_from: str = Query(""), date_to: str = Query(""),
    payment_mode: str = Query(""), paid_by: str = Query(""),
    page: int = Query(1, ge=1), limit: int = Query(15, ge=1, le=100),
    user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))
):
    bid = user.get("branch_id")
    query = {"branch_id": bid, "is_deleted": False}
    if date_from:
        query["date"] = {"$gte": date_from}
    if date_to:
        query.setdefault("date", {})["$lte"] = date_to
    if payment_mode:
        query["payment_mode"] = payment_mode
    if paid_by:
        query["payment_done_by"] = paid_by
    skip = (page - 1) * limit
    total = await db.branch_expenses.count_documents(query)
    expenses = await db.branch_expenses.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(limit).to_list(limit)
    return {"expenses": expenses, "total": total, "page": page, "limit": limit}

@api_router.get("/branch/expenses/export")
async def export_expenses(
    date_from: str = Query(""), date_to: str = Query(""),
    payment_mode: str = Query(""), paid_by: str = Query(""),
    user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))
):
    bid = user.get("branch_id")
    query = {"branch_id": bid, "is_deleted": False}
    if date_from:
        query["date"] = {"$gte": date_from}
    if date_to:
        query.setdefault("date", {})["$lte"] = date_to
    if payment_mode:
        query["payment_mode"] = payment_mode
    if paid_by:
        query["payment_done_by"] = paid_by
    expenses = await db.branch_expenses.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"
    headers = ["Date", "Expense For", "Amount (₹)", "Paid By", "Mode of Payment", "Note", "Created By", "Created At"]
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="F8FAFC")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for e in expenses:
        ws.append([
            e.get("date", ""),
            e.get("expense_for", ""),
            e.get("amount", 0),
            e.get("payment_done_by_name", ""),
            e.get("payment_mode", ""),
            e.get("note", ""),
            e.get("created_by_name", ""),
            e.get("created_at", "")[:10] if e.get("created_at") else ""
        ])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 14
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await log_audit(user["id"], bid, "EXPORT_EXPENSES", "", f"Exported {len(expenses)} expense records")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=expenses_export.xlsx"}
    )

@api_router.put("/branch/expenses/{expense_id}")
async def update_expense(expense_id: str, data: ExpenseUpdate, user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    expense = await db.branch_expenses.find_one({"id": expense_id, "branch_id": bid, "is_deleted": False})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    # Permission check
    if user["role"] != "BRANCH_HEAD" and expense.get("created_by_id") != user["id"]:
        raise HTTPException(status_code=403, detail="You can only edit your own expenses")
    if data.amount is not None and data.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be positive")
    update_fields = {k: v for k, v in data.model_dump().items() if v is not None}
    now = datetime.now(timezone.utc).isoformat()
    update_fields["last_updated_by"] = user["id"]
    update_fields["last_updated_at"] = now
    await db.branch_expenses.update_one({"id": expense_id}, {"$set": update_fields})
    await log_audit(user["id"], bid, "UPDATE_EXPENSE", expense_id,
                    f"Updated expense: {expense.get('expense_for', '')}")
    updated = await db.branch_expenses.find_one({"id": expense_id}, {"_id": 0})
    return updated

@api_router.delete("/branch/expenses/{expense_id}")
async def delete_expense(expense_id: str, user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    bid = user.get("branch_id")
    expense = await db.branch_expenses.find_one({"id": expense_id, "branch_id": bid, "is_deleted": False})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    # Permission check
    if user["role"] != "BRANCH_HEAD" and expense.get("created_by_id") != user["id"]:
        raise HTTPException(status_code=403, detail="You can only delete your own expenses")
    now = datetime.now(timezone.utc).isoformat()
    await db.branch_expenses.update_one(
        {"id": expense_id},
        {"$set": {"is_deleted": True, "deleted_at": now, "last_updated_by": user["id"], "last_updated_at": now}}
    )
    await log_audit(user["id"], bid, "DELETE_EXPENSE", expense_id,
                    f"Deleted expense: {expense.get('expense_for', '')} ₹{expense.get('amount', 0)}")
    return {"message": "Expense deleted"}

@api_router.post("/branch/expenses/upload-attachment")
async def upload_expense_attachment(file: UploadFile = File(...), user: dict = Depends(require_role("BRANCH_HEAD", "STAFF"))):
    allowed = {".pdf", ".jpg", ".jpeg", ".png", ".webp"}
    ext = Path(file.filename).suffix.lower() if file.filename else ".pdf"
    if ext not in allowed:
        raise HTTPException(status_code=400, detail="Only PDF, JPG, PNG, WEBP files allowed")
    filename = f"expense_{uuid.uuid4()}{ext}"
    filepath = UPLOAD_DIR / filename
    with open(filepath, "wb") as f:
        content = await file.read()
        f.write(content)
    return {"file_url": f"/api/files/{filename}", "filename": file.filename}

# ==================== POLL ENDPOINTS ====================

class PollVoteInput(BaseModel):
    name: str
    score_range: str

class PollCommentInput(BaseModel):
    name: str
    comment: str

@api_router.post("/poll/vote")
async def submit_poll_vote(vote: PollVoteInput):
    vote_data = vote.model_dump()
    vote_data["id"] = str(uuid.uuid4())
    vote_data["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.poll_votes.insert_one(vote_data)
    return {"message": "Vote submitted successfully"}

@api_router.get("/poll/results")
async def get_poll_results():
    pipeline = [
        {"$group": {"_id": "$score_range", "count": {"$sum": 1}}}
    ]
    results = await db.poll_votes.aggregate(pipeline).to_list(None)
    
    formatted = {item["_id"]: item["count"] for item in results}
    total_votes = await db.poll_votes.count_documents({})
    return {"results": formatted, "total": total_votes}

@api_router.post("/poll/comment")
async def submit_poll_comment(comment: PollCommentInput):
    comment_data = comment.model_dump()
    comment_data["id"] = str(uuid.uuid4())
    comment_data["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.poll_comments.insert_one(comment_data)
    return {"message": "Comment added successfully"}

@api_router.get("/poll/comments")
async def get_poll_comments():
    comments = await db.poll_comments.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"comments": comments}

# ==================== STARTUP ====================

# ==================== USER MANAGEMENT (ADMIN ONLY) ====================

@api_router.get("/admin/users")
async def get_all_users(current_user: Dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    raw_users = await db.users.find({}, {"password": 0}).to_list(length=1000)
    
    # Convert ObjectId to string for JSON serialization
    users = []
    for u in raw_users:
        u["_id"] = str(u["_id"])
        users.append(u)
    
    return users

@api_router.put("/admin/users/{user_id}")
async def admin_update_user(user_id: str, data: Dict, current_user: Dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    update_data = {}
    if "name" in data: update_data["name"] = data["name"]
    if "email" in data: update_data["email"] = data["email"]
    if "role" in data: update_data["role"] = data["role"]
    if "is_active" in data: update_data["is_active"] = data["is_active"]
    if "password" in data and data["password"]:
        update_data["password"] = hash_password(data["password"])
        
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    result = await db.users.update_one({"id": user_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
        
    return {"message": "User updated successfully"}

@api_router.delete("/admin/users/{user_id}")
async def admin_delete_user(user_id: str, current_user: Dict = Depends(get_current_user)):
    if current_user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Don't allow deleting self
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="You cannot delete your own admin account")
        
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
        
    return {"message": "User deleted successfully"}

@app.on_event("startup")
async def startup():
    logging.info("🚀 Starting non-blocking database initialization...")
    try:
        # Create indexes only if connection is ready
        await db.users.create_index("email", unique=True)
        await db.users.create_index("id", unique=True)
        await db.branches.create_index("id", unique=True)
        await db.student_profiles.create_index("id", unique=True)
        await db.student_profiles.create_index("user_id")
        await db.student_profiles.create_index("branch_id")
        await db.documents.create_index("student_id")
        await db.audit_logs.create_index("branch_id")
        await db.verification_logs.create_index("student_id")
        await db.payments.create_index("student_id")
        await db.chat_messages.create_index("student_id")
        await db.chat_messages.create_index([("student_id", 1), ("created_at", 1)])
        await db.notifications.create_index("user_id")
        await db.notifications.create_index([("user_id", 1), ("is_read", 1)])
        await db.branch_expenses.create_index("id", unique=True)
        await db.branch_expenses.create_index("branch_id")
        await db.branch_expenses.create_index([("branch_id", 1), ("date", -1)])
        await db.registration_history.create_index([("student_id", 1), ("version", -1)])
        
        # Admin Seeding
        admin_emails = ["admin@ame.com", "virus392k@gmail.com"]
        for email in admin_emails:
            existing = await db.users.find_one({"email": email})
            if not existing:
                await db.users.insert_one({
                    "id": str(uuid.uuid4()), 
                    "email": email,
                    "password": hash_password("admin123" if "admin" in email else "Virus@392k"), 
                    "name": "Super Admin",
                    "role": "ADMIN", 
                    "branch_id": None, 
                    "is_active": True,
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
                logging.info(f"✅ Seeded admin: {email}")

        # Public Admin Seeding
        public_admin_email = "vikrant39@gmail.com"
        public_admin_phone = "3920003920"
        existing_pa = await db.public_admins.find_one({"email": public_admin_email})
        if not existing_pa:
            await db.public_admins.insert_one({
                "id": str(uuid.uuid4()),
                "email": public_admin_email,
                "phone": public_admin_phone,
                "password": hash_password("Virus@392k"),
                "name": "Vikrant (Master Admin)",
                "role": "PUBLIC_ADMIN",
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            logging.info(f"✅ Seeded public admin: {public_admin_email}")

        # Public Admin Indexes
        await db.public_admins.create_index("id", unique=True)
        await db.public_admins.create_index("email", unique=True)
        await db.public_admins.create_index("phone", unique=True)
        await db.public_chats.create_index([("student_id", 1), ("created_at", 1)])
        await db.public_chats.create_index([("student_id", 1), ("sender_role", 1), ("read", 1)])
        await db.public_audit_logs.create_index([("created_at", -1)])
        await db.public_broadcasts.create_index([("created_at", -1)])

    except Exception as e:
        logging.error(f"⚠️ DATABASE INITIALIZATION FAILED (Handshake issue?): {e}")
        logging.info("Continuing startup to satisfy cloud health checks...")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

app.include_router(api_router, prefix="/api")
# ---- FINAL CORS CONFIGURATION ----
cors_origins_str = os.environ.get('CORS_ORIGINS', '*')
if cors_origins_str == '*':
    # If wildcard is used with credentials, we must use allow_origin_regex or a specific list.
    # For security and compatibility with credentials, it's better to list the domains or use * if credentials not needed.
    # Here we allow all origins by echoing the Origin header if credentials are True.
    # Note: Using ["*"] with allow_credentials=True is not allowed by browsers.
    cors_origins = ["*"]
else:
    cors_origins = [o.strip() for o in cors_origins_str.split(',') if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins if cors_origins_str != '*' else [],
    allow_origin_regex=".*" if cors_origins_str == '*' else None,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("server:app", host="0.0.0.0", port=port, reload=False)
